import streamlit as st
import psycopg
from sqlalchemy import create_engine
import pandas as pd
import io
import pdfplumber
import re
from datetime import datetime
import calendar
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import gc
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="Gestão de Energia - DAE", layout="wide", page_icon="⚡")

# --- 1. SISTEMA DE LOGIN DAE ---
def check_password():
    """Valida se o usuário e senha digitados batem com os dados do Secrets."""
    def password_entered():
        usuario_digitado = st.session_state["username"]
        senha_digitada = st.session_state["password"]
        
        # Verifica se o usuário existe no secrets e se a senha confere
        if usuario_digitado in st.secrets["usuarios"] and senha_digitada == st.secrets["usuarios"][usuario_digitado]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Apaga a senha da memória por segurança
        else:
            # Se errar, aí sim o status vira Falso
            st.session_state["password_correct"] = False

    # Retorna True se o usuário já estiver logado com sucesso
    if st.session_state.get("password_correct"):
        return True

    # Se chegou até aqui, é porque não está logado. Desenha a tela:
    col1, col2, col3 = st.columns([2, 1.5, 2])
    with col2:
        st.write("")
        st.write("")
        st.write("") # Espaçadores
        
        # Carrega a sua logo real e centraliza usando colunas internas
        col_img1, col_img2, col_img3 = st.columns([1, 1.5, 1])
        with col_img2:
            try:
                st.image("logo_DAE.png", use_container_width=True)
            except:
                pass # Caso o nome da logo esteja diferente, ele não quebra a tela
        
        st.markdown("<h2 style='text-align: center; color: #0055A5;'>🔐 Acesso Restrito</h2>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center;'>Insira suas credenciais para acessar o Sistema de Faturas de Energia.</p>", unsafe_allow_html=True)
        
        st.text_input("👤 Usuário DAE", key="username")
        st.text_input("🔑 Senha", type="password", key="password")
        
        st.button("🚀 Entrar no Sistema", type="primary", on_click=password_entered, use_container_width=True)
        
        # O erro só aparece se o status for EXPLICITAMENTE False (após tentativa de login)
        if st.session_state.get("password_correct") == False:
            st.error("🚨 Usuário ou senha incorretos. Tente novamente.")
            
    return False # Bloqueia o acesso ao resto do código

# --- 2. A "CATRACA" DO STREAMLIT ---
# Se a função retornar falso (não logado), o st.stop() mata o código na hora e esconde o painel
if not check_password():
    st.stop()

# --- OCULTAR ELEMENTOS PADRÃO DO STREAMLIT ---
esconder_botoes = """
    <style>
    /* Oculta o menu hambúrguer no canto superior direito */
    #MainMenu {visibility: hidden;}
    
    /* Oculta o rodapé "Made with Streamlit" */
    footer {visibility: hidden;}
    
    /* Oculta o botão de "Deploy" se ele estiver aparecendo */
    .stAppDeployButton {display:none;}
    </style>
"""
st.markdown(esconder_botoes, unsafe_allow_html=True)

# --- CONEXÃO SEGURA COM O BANCO DE DADOS ---
# Se for rodar localmente antes de por na nuvem, crie uma pasta .streamlit e um arquivo secrets.toml
try:
    DATABASE_URL = st.secrets["DATABASE_URL"]
except FileNotFoundError:
    st.error("Erro: Arquivo secrets.toml não encontrado ou DATABASE_URL não configurada no Streamlit Cloud.")
    st.stop()

def obter_conexao():
    return psycopg.connect(DATABASE_URL, autocommit=True, prepare_threshold=None)

# --- CABEÇALHO COM LOGOTIPO ---
col_logo, col_titulo, _ = st.columns([0.6, 5, 1])

with col_logo:
    try:
        st.image("logo_DAE.png", use_container_width=True)
    except:
        pass

with col_titulo:
    st.subheader("⚡ Sistema de Faturas de Energia - DAE Bauru")

# --- 1. BANCO DE DADOS: CRIAÇÃO E PRÉ-CADASTRO ---
def inicializar_banco():
    conexao = obter_conexao()
    cursor = conexao.cursor()
    
    # 1. Cria a tabela principal de faturas (Mudança de AUTOINCREMENT para SERIAL no Postgres)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS faturas_cpfl (
            id SERIAL PRIMARY KEY,
            classificacao TEXT, unidade_consumidora TEXT, nome_unidade TEXT, atividade TEXT,
            periodo_leitura_inicio TEXT, periodo_leitura_fim TEXT, data_proxima_leitura TEXT,
            mes_referencia TEXT, data_vencimento TEXT,
            demanda_contratada_ponta REAL, demanda_contratada_fponta REAL,
            consumo_ponta REAL, tarifa_aneel_cons_ponta_tusd REAL, tarifa_trib_cons_ponta_tusd REAL, valor_cons_ponta_tusd REAL,
            consumo_fora_ponta REAL, tarifa_aneel_cons_fponta_tusd REAL, tarifa_trib_cons_fponta_tusd REAL, valor_cons_fponta_tusd REAL,
            tarifa_aneel_cons_ponta_te REAL, tarifa_trib_cons_ponta_te REAL, valor_cons_ponta_te REAL,
            tarifa_aneel_cons_fponta_te REAL, tarifa_trib_cons_fponta_te REAL, valor_cons_fponta_te REAL,
            tipo_bandeira TEXT, adicional_bandeira REAL,
            demanda_registrada_ponta REAL, tarifa_aneel_dem_ponta REAL, tarifa_trib_dem_ponta REAL, valor_dem_ponta REAL,
            demanda_isenta_ponta REAL, tarifa_aneel_dem_isenta_ponta REAL, tarifa_trib_dem_isenta_ponta REAL, valor_dem_isenta_ponta REAL,
            demanda_registrada_fora_ponta REAL, tarifa_aneel_dem_fponta REAL, tarifa_trib_dem_fponta REAL, valor_dem_fponta REAL,
            demanda_isenta_fora_ponta REAL, tarifa_aneel_dem_isenta_fponta REAL, tarifa_trib_dem_isenta_fponta REAL, valor_dem_isenta_fponta REAL,
            consumo_reativo_ponta REAL, tarifa_aneel_cons_reativo_ponta REAL, tarifa_trib_cons_reativo_ponta REAL, valor_cons_reativo_ponta REAL,
            consumo_reativo_fora_ponta REAL, tarifa_aneel_cons_reativo_fponta REAL, tarifa_trib_cons_reativo_fponta REAL, valor_cons_reativo_fponta REAL,
            demanda_ultrapassagem_ponta REAL, tarifa_aneel_dem_ultrap_ponta REAL, tarifa_trib_dem_ultrap_ponta REAL, valor_dem_ultrap_ponta REAL,
            demanda_ultrapassagem_fora_ponta REAL, tarifa_aneel_dem_ultrap_fponta REAL, tarifa_trib_dem_ultrap_fponta REAL, valor_dem_ultrap_fponta REAL,
            demanda_reativa_ponta REAL, tarifa_aneel_dem_reativa_ponta REAL, tarifa_trib_dem_reativa_ponta REAL, valor_dem_reativa_ponta REAL,
            demanda_reativa_fora_ponta REAL, tarifa_aneel_dem_reativa_fponta REAL, tarifa_trib_dem_reativa_fponta REAL, valor_dem_reativa_fponta REAL,
            cip REAL, retencao_consumo_irrf REAL, retencao_demanda_irrf REAL,
            valor_total_pis REAL, valor_total_cofins REAL, valor_total_icms REAL, valor_total_fatura REAL,
            data_insercao TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 2. Cria a tabela de Cadastro de UCs (Agora com STATUS e DIA VENCIMENTO)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cadastro_uc (
            unidade_consumidora TEXT PRIMARY KEY, nome_unidade TEXT, 
            atividade TEXT, classificacao TEXT, demanda_contratada_ponta REAL, 
            demanda_contratada_fponta REAL, status TEXT DEFAULT 'ATIVA',
            dia_vencimento INTEGER DEFAULT 10
        )
    ''')
    # Comando de segurança para adicionar colunas em bancos que já existem
    try:
        cursor.execute('''ALTER TABLE cadastro_uc ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'ATIVA';''')
        cursor.execute('''ALTER TABLE cadastro_uc ADD COLUMN IF NOT EXISTS dia_vencimento INTEGER DEFAULT 10;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS subtotal_fatura REAL DEFAULT 0.0;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS consumo_energia_acl_kwh DOUBLE PRECISION DEFAULT 0.0;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS tarifa_energia_acl DOUBLE PRECISION DEFAULT 0.0;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS valor_energia_acl DOUBLE PRECISION DEFAULT 0.0;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS valor_total_acl DOUBLE PRECISION DEFAULT 0.0;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS data_vencimento_acl TEXT;''')
        cursor.execute('''ALTER TABLE cadastro_uc ADD COLUMN IF NOT EXISTS uc_cemig TEXT;''')
        cursor.execute('''ALTER TABLE cadastro_uc ADD COLUMN IF NOT EXISTS uc_antiga TEXT;''')
        cursor.execute('''ALTER TABLE faturas_cpfl ADD COLUMN IF NOT EXISTS uc_original TEXT;''')
    except Exception as e:
        print("Aviso ao atualizar estrutura do banco:", e)
        
    # 3. Cria a tabela de Histórico de Envios para o Financeiro
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS historico_financeiro (
            id SERIAL PRIMARY KEY, unidade_consumidora TEXT,
            mes_referencia TEXT, valor_fatura REAL, vencimento TEXT,
            data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    try:
        for tabela in ['faturas_cpfl', 'cadastro_uc', 'historico_financeiro']:
            cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{tabela}' AND data_type = 'real';")
            colunas_baixa_precisao = cursor.fetchall()
            for col in colunas_baixa_precisao:
                cursor.execute(f"ALTER TABLE {tabela} ALTER COLUMN {col[0]} TYPE DOUBLE PRECISION;")
    except Exception as e:
        print("Aviso ao atualizar precisão do banco:", e)
    
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_faturas_uc_mes ON faturas_cpfl (unidade_consumidora, mes_referencia);")
    
    conexao.commit()
    conexao.close()

inicializar_banco()

# --- 2. FUNÇÕES DE EXTRAÇÃO DE PDF E MANIPULAÇÃO DE DADOS ---
@st.cache_data(show_spinner="Carregando e processando banco de dados...", ttl=600, max_entries=2)
def carregar_dados():
    # Usando SQLAlchemy para facilitar a vida do Pandas ao ler do Postgres
    url_sqlalchemy = DATABASE_URL.replace("postgresql://", "postgresql+psycopg://")
    engine = create_engine(url_sqlalchemy)
    df = pd.read_sql_query("SELECT * FROM faturas_cpfl", engine)
    
    if df.empty:
        return df
        
    dicionario_nomes = {
        'classificacao': 'Classificação', 'unidade_consumidora': 'UC', 'nome_unidade': 'Nome da Unidade',
        'atividade': 'Atividade', 'periodo_leitura_inicio': 'Leitura Anterior', 'periodo_leitura_fim': 'Leitura Atual',
        'data_proxima_leitura': 'Próxima Leitura', 'mes_referencia': 'Mês Referência', 'data_vencimento': 'Vencimento CPFL',
        'demanda_contratada_ponta': 'Dem. Contr. Ponta', 'demanda_contratada_fponta': 'Dem. Contr. F.Ponta',
        'consumo_ponta': 'Consumo Ponta', 'tarifa_aneel_cons_ponta_tusd': 'Tarifa Cons. Ponta TUSD',
        'tarifa_trib_cons_ponta_tusd': 'Tarifa Trib. Cons. Ponta TUSD', 'valor_cons_ponta_tusd': 'Valor Cons. Ponta TUSD',
        'consumo_fora_ponta': 'Consumo F.Ponta', 'tarifa_aneel_cons_fponta_tusd': 'Tarifa Cons. F.Ponta TUSD',
        'tarifa_trib_cons_fponta_tusd': 'Tarifa Trib. Cons. F.Ponta TUSD', 'valor_cons_fponta_tusd': 'Valor Cons. F.Ponta TUSD',
        'tarifa_aneel_cons_ponta_te': 'Tarifa Cons. Ponta TE', 'tarifa_trib_cons_ponta_te': 'Tarifa Trib. Cons. Ponta TE',
        'valor_cons_ponta_te': 'Valor Cons. Ponta TE', 'tarifa_aneel_cons_fponta_te': 'Tarifa Cons. F.Ponta TE',
        'tarifa_trib_cons_fponta_te': 'Tarifa Trib. Cons. F.Ponta TE', 'valor_cons_fponta_te': 'Valor Cons. F.Ponta TE',
        'tipo_bandeira': 'Bandeira', 'adicional_bandeira': 'Adicional Bandeira', 'demanda_registrada_ponta': 'Dem. Reg. Ponta',
        'tarifa_aneel_dem_ponta': 'Tarifa Dem. Ponta', 'tarifa_trib_dem_ponta': 'Tarifa Trib. Dem. Ponta',
        'valor_dem_ponta': 'Valor Dem. Ponta', 'demanda_isenta_ponta': 'Dem. Isenta Ponta',
        'tarifa_aneel_dem_isenta_ponta': 'Tarifa Dem. Isenta Ponta', 'tarifa_trib_dem_isenta_ponta': 'Tarifa Trib. Dem. Isenta Ponta',
        'valor_dem_isenta_ponta': 'Valor Dem. Isenta Ponta', 'demanda_registrada_fora_ponta': 'Dem. Reg. F.Ponta',
        'tarifa_aneel_dem_fponta': 'Tarifa Dem. F.Ponta', 'tarifa_trib_dem_fponta': 'Tarifa Trib. Dem. F.Ponta',
        'valor_dem_fponta': 'Valor Dem. F.Ponta', 'demanda_isenta_fora_ponta': 'Dem. Isenta F.Ponta',
        'tarifa_aneel_dem_isenta_fponta': 'Tarifa Dem. Isenta F.Ponta', 'tarifa_trib_dem_isenta_fponta': 'Tarifa Trib. Dem. Isenta F.Ponta',
        'valor_dem_isenta_fponta': 'Valor Dem. Isenta F.Ponta', 'consumo_reativo_ponta': 'Cons. Reat. Ponta',
        'tarifa_aneel_cons_reativo_ponta': 'Tarifa Cons. Reat. Ponta', 'tarifa_trib_cons_reativo_ponta': 'Tarifa Trib. Cons. Reat. Ponta',
        'valor_cons_reativo_ponta': 'Valor Cons. Reat. Ponta', 'consumo_reativo_fora_ponta': 'Cons. Reat. F.Ponta',
        'tarifa_aneel_cons_reativo_fponta': 'Tarifa Cons. Reat. F.Ponta', 'tarifa_trib_cons_reativo_fponta': 'Tarifa Trib. Cons. Reat. F.Ponta',
        'valor_cons_reativo_fponta': 'Valor Cons. Reat. F.Ponta', 'demanda_ultrapassagem_ponta': 'Dem. Ultrap. Ponta',
        'tarifa_aneel_dem_ultrap_ponta': 'Tarifa Dem. Ultrap. Ponta', 'tarifa_trib_dem_ultrap_ponta': 'Tarifa Trib. Dem. Ultrap. Ponta',
        'valor_dem_ultrap_ponta': 'Valor Dem. Ultrap. Ponta', 'demanda_ultrapassagem_fora_ponta': 'Dem. Ultrap. F.Ponta',
        'tarifa_aneel_dem_ultrap_fponta': 'Tarifa Dem. Ultrap. F.Ponta', 'tarifa_trib_dem_ultrap_fponta': 'Tarifa Trib. Dem. Ultrap. F.Ponta',
        'valor_dem_ultrap_fponta': 'Valor Dem. Ultrap. F.Ponta', 'demanda_reativa_ponta': 'Dem. Reat. Ponta',
        'tarifa_aneel_dem_reativa_ponta': 'Tarifa Dem. Reat. Ponta', 'tarifa_trib_dem_reativa_ponta': 'Tarifa Trib. Dem. Reat. Ponta',
        'valor_dem_reativa_ponta': 'Valor Dem. Reat. Ponta', 'demanda_reativa_fora_ponta': 'Dem. Reat. F.Ponta',
        'tarifa_aneel_dem_reativa_fponta': 'Tarifa Dem. Reat. F.Ponta', 'tarifa_trib_dem_reativa_fponta': 'Tarifa Trib. Dem. Reat. F.Ponta',
        'valor_dem_reativa_fponta': 'Valor Dem. Reat. F.Ponta', 'subtotal_fatura': 'Subtotal PDF', 'cip': 'CIP', 'retencao_consumo_irrf': 'Retenção Cons. IRRF',
        'retencao_demanda_irrf': 'Retenção Dem. IRRF', 'valor_total_pis': 'Valor PIS', 'valor_total_cofins': 'Valor COFINS',
        'valor_total_icms': 'Valor ICMS', 'valor_total_fatura': 'Valor Total Fatura', 'data_insercao': 'Data Cadastro',
        'data_vencimento_acl': 'Vencimento ACL', 'consumo_energia_acl_kwh': 'Consumo Energia ACL (kWh)',
        'tarifa_energia_acl': 'Tarifa Energia ACL (R$/kWh)', 'valor_energia_acl': 'Valor Energia ACL (R$)',
        'valor_total_acl': 'Valor Total ACL (R$)'
    }
    
    df = df.rename(columns=dicionario_nomes)
    df['Total Consumo'] = df['Consumo Ponta'] + df['Consumo F.Ponta']
    df['Valor Total Consumo'] = df['Valor Cons. Ponta TUSD'] + df['Valor Cons. Ponta TE'] + df['Valor Cons. F.Ponta TUSD'] + df['Valor Cons. F.Ponta TE']
    df['Valor Total Dem. Isenta'] = df['Valor Dem. Isenta Ponta'] + df['Valor Dem. Isenta F.Ponta']
    df['Valor Total Dem. Ultrap.'] = df['Valor Dem. Ultrap. Ponta'] + df['Valor Dem. Ultrap. F.Ponta']
    df['Valor Total Desv. Dem.'] = df['Valor Dem. Isenta F.Ponta']+df['Valor Dem. Isenta Ponta']+df['Valor Dem. Ultrap. F.Ponta']+df['Valor Dem. Ultrap. Ponta']
    df['Total Cons. Reat.'] = df['Cons. Reat. Ponta'] + df['Cons. Reat. F.Ponta']
    df['Valor Total Cons. Reat.'] = df['Valor Cons. Reat. Ponta'] + df['Valor Cons. Reat. F.Ponta']
    df['Valor Total Dem. Reat.'] = df['Valor Dem. Reat. Ponta'] + df['Valor Dem. Reat. F.Ponta']
    df['Valor Total Dem.'] = df['Valor Dem. Ponta'] + df['Valor Dem. F.Ponta']
    df['Valor Total Reativo'] = df['Valor Total Cons. Reat.'] + df['Valor Total Dem. Reat.']

    mes_map = {'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04', 'MAI': '05', 'JUN': '06', 
               'JUL': '07', 'AGO': '08', 'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'}
    
    def converter_para_data_real(mes_str):
        try:
            m, y = str(mes_str).split('/')
            return pd.to_datetime(f"{y}-{mes_map[m.upper()]}-01")
        except:
            return pd.NaT

    df['Data Referência Oculta'] = df['Mês Referência'].apply(converter_para_data_real)
    df = df.sort_values(by=['Data Referência Oculta', 'UC'], ascending=[False, True])

    ordem_colunas = [
        'id', 'Data Referência Oculta', 'UC', 'Nome da Unidade', 'Atividade', 'Classificação', 'Mês Referência', 'Vencimento CPFL', 'Vencimento ACL',
        'Leitura Anterior', 'Leitura Atual', 'Próxima Leitura', 'Consumo Energia ACL (kWh)', 'Tarifa Energia ACL (R$/kWh)', 'Valor Energia ACL (R$)', 'Valor Total ACL (R$)',
        'Consumo Ponta', 'Tarifa Cons. Ponta TUSD', 'Tarifa Trib. Cons. Ponta TUSD', 'Valor Cons. Ponta TUSD', 
        'Tarifa Cons. Ponta TE', 'Tarifa Trib. Cons. Ponta TE', 'Valor Cons. Ponta TE', 'Consumo F.Ponta', 'Tarifa Cons. F.Ponta TUSD', 'Tarifa Trib. Cons. F.Ponta TUSD', 'Valor Cons. F.Ponta TUSD', 
        'Tarifa Cons. F.Ponta TE', 'Tarifa Trib. Cons. F.Ponta TE', 'Valor Cons. F.Ponta TE', 'Bandeira', 'Adicional Bandeira', 
        'Dem. Contr. Ponta', 'Dem. Reg. Ponta', 'Tarifa Dem. Ponta', 'Tarifa Trib. Dem. Ponta', 'Valor Dem. Ponta', 
        'Dem. Isenta Ponta', 'Tarifa Dem. Isenta Ponta', 'Tarifa Trib. Dem. Isenta Ponta', 'Valor Dem. Isenta Ponta', 
        'Dem. Ultrap. Ponta', 'Tarifa Dem. Ultrap. Ponta', 'Tarifa Trib. Dem. Ultrap. Ponta', 'Valor Dem. Ultrap. Ponta', 
        'Dem. Contr. F.Ponta', 'Dem. Reg. F.Ponta', 'Tarifa Dem. F.Ponta', 'Tarifa Trib. Dem. F.Ponta', 'Valor Dem. F.Ponta', 
        'Dem. Isenta F.Ponta', 'Tarifa Dem. Isenta F.Ponta', 'Tarifa Trib. Dem. Isenta F.Ponta', 'Valor Dem. Isenta F.Ponta', 
        'Dem. Ultrap. F.Ponta', 'Tarifa Dem. Ultrap. F.Ponta', 'Tarifa Trib. Dem. Ultrap. F.Ponta', 'Valor Dem. Ultrap. F.Ponta', 
        'Cons. Reat. Ponta', 'Tarifa Cons. Reat. Ponta', 'Tarifa Trib. Cons. Reat. Ponta', 'Valor Cons. Reat. Ponta', 
        'Cons. Reat. F.Ponta', 'Tarifa Cons. Reat. F.Ponta', 'Tarifa Trib. Cons. Reat. F.Ponta', 'Valor Cons. Reat. F.Ponta', 
        'Dem. Reat. Ponta', 'Tarifa Dem. Reat. Ponta', 'Tarifa Trib. Dem. Reat. Ponta', 'Valor Dem. Reat. Ponta', 
        'Dem. Reat. F.Ponta', 'Tarifa Dem. Reat. F.Ponta', 'Tarifa Trib. Dem. Reat. F.Ponta', 'Valor Dem. Reat. F.Ponta', 
        'Subtotal PDF', 'CIP', 'Retenção Cons. IRRF', 'Retenção Dem. IRRF', 'Valor PIS', 'Valor COFINS', 'Valor ICMS', 
        'Total Consumo', 'Valor Total Consumo', 'Valor Total Dem.', 'Valor Total Dem. Isenta', 'Valor Total Dem. Ultrap.', 
        'Valor Total Desv. Dem.', 'Total Cons. Reat.', 'Valor Total Cons. Reat.', 'Valor Total Dem. Reat.', 'Valor Total Reativo',
        'Valor Total Fatura', 'Data Cadastro'
    ]
    
    colunas_categoria = ['UC', 'Nome da Unidade', 'Atividade', 'Classificação', 'Mês Referência', 'Bandeira']
    for col in colunas_categoria:
        if col in df.columns:
            df[col] = df[col].astype('category')

    colunas_finais = [c for c in ordem_colunas if c in df.columns]
    return df[colunas_finais]

def limpar_numero(texto_numero):
    if not texto_numero: return 0.0
    return float(texto_numero.replace('.', '').replace(',', '.'))

def extrair_texto_regex(padrao, texto, grupo=1, padrao_falha=""):
    res = re.search(padrao, texto, re.IGNORECASE | re.MULTILINE)
    return res.group(grupo).strip() if res else padrao_falha

def extrair_valor_regex(padrao, texto, grupo=1):
    res = re.search(padrao, texto, re.IGNORECASE | re.MULTILINE)
    return limpar_numero(res.group(grupo)) if res else 0.0

def processar_pdf(arquivo_pdf):
    with pdfplumber.open(arquivo_pdf) as pdf:
        texto = ""
        for i in range(min(2, len(pdf.pages))):
            texto_pagina = pdf.pages[i].extract_text()
            if texto_pagina:
                texto += texto_pagina + "\n"
    
    chaves_numericas = [
        'demanda_contratada_ponta', 'demanda_contratada_fponta',
        'consumo_ponta', 'tarifa_aneel_cons_ponta_tusd', 'tarifa_trib_cons_ponta_tusd', 'valor_cons_ponta_tusd',
        'consumo_fora_ponta', 'tarifa_aneel_cons_fponta_tusd', 'tarifa_trib_cons_fponta_tusd', 'valor_cons_fponta_tusd',
        'tarifa_aneel_cons_ponta_te', 'tarifa_trib_cons_ponta_te', 'valor_cons_ponta_te',
        'tarifa_aneel_cons_fponta_te', 'tarifa_trib_cons_fponta_te', 'valor_cons_fponta_te',
        'demanda_registrada_ponta', 'tarifa_aneel_dem_ponta', 'tarifa_trib_dem_ponta', 'valor_dem_ponta',
        'demanda_isenta_ponta', 'tarifa_aneel_dem_isenta_ponta', 'tarifa_trib_dem_isenta_ponta', 'valor_dem_isenta_ponta',
        'demanda_registrada_fora_ponta', 'tarifa_aneel_dem_fponta', 'tarifa_trib_dem_fponta', 'valor_dem_fponta',
        'demanda_isenta_fora_ponta', 'tarifa_aneel_dem_isenta_fponta', 'tarifa_trib_dem_isenta_fponta', 'valor_dem_isenta_fponta',
        'consumo_reativo_ponta', 'tarifa_aneel_cons_reativo_ponta', 'tarifa_trib_cons_reativo_ponta', 'valor_cons_reativo_ponta',
        'consumo_reativo_fora_ponta', 'tarifa_aneel_cons_reativo_fponta', 'tarifa_trib_cons_reativo_fponta', 'valor_cons_reativo_fponta',
        'demanda_ultrapassagem_ponta', 'tarifa_aneel_dem_ultrap_ponta', 'tarifa_trib_dem_ultrap_ponta', 'valor_dem_ultrap_ponta',
        'demanda_ultrapassagem_fora_ponta', 'tarifa_aneel_dem_ultrap_fponta', 'tarifa_trib_dem_ultrap_fponta', 'valor_dem_ultrap_fponta',
        'demanda_reativa_ponta', 'tarifa_aneel_dem_reativa_ponta', 'tarifa_trib_dem_reativa_ponta', 'valor_dem_reativa_ponta',
        'demanda_reativa_fora_ponta', 'tarifa_aneel_dem_reativa_fponta', 'tarifa_trib_dem_reativa_fponta', 'valor_dem_reativa_fponta', 'subtotal_fatura'
    ]
    dados = {k: 0.0 for k in chaves_numericas}
    
    classificacao_bruta = extrair_texto_regex(r"Classificação:\s*(.*?)(?:\s+Serviço|\s+Tipo|\n)", texto)
    if "B3" in classificacao_bruta.upper():
        dados['classificacao'] = "Convencional B3"
    elif "VERDE" in classificacao_bruta.upper():
        dados['classificacao'] = "Tarifa Verde-A4"
    elif "AZUL" in classificacao_bruta.upper():
        dados['classificacao'] = "Tarifa Azul-A4"
    else:
        dados['classificacao'] = classificacao_bruta

    dados['unidade_consumidora'] = extrair_texto_regex(r"(\d{7,12})\s+\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+\d{1,3}", texto)
    
    conexao_pdf = obter_conexao()
    c_pdf = conexao_pdf.cursor()
    c_pdf.execute("SELECT nome_unidade, atividade, demanda_contratada_ponta, demanda_contratada_fponta FROM cadastro_uc WHERE unidade_consumidora = %s", (dados['unidade_consumidora'],))
    res_uc = c_pdf.fetchone()
    conexao_pdf.close()
    
    dados['nome_unidade'] = res_uc[0] if res_uc else "Não Cadastrada"
    dados['atividade'] = res_uc[1] if res_uc else "Administrativa" 
    dados['demanda_contratada_ponta'] = res_uc[2] if res_uc else 0.0
    dados['demanda_contratada_fponta'] = res_uc[3] if res_uc else 0.0
    
    dc_p_pdf = extrair_valor_regex(r"Demanda P\.? kW\s+([\d\.,]+)", texto)
    if dc_p_pdf > 0: dados['demanda_contratada_ponta'] = dc_p_pdf
    
    dc_fp_pdf = extrair_valor_regex(r"Demanda FP\.? kW\s+([\d\.,]+)", texto)
    if dc_fp_pdf > 0: dados['demanda_contratada_fponta'] = dc_fp_pdf
    
    if "Verde" in dados['classificacao']:
        dc_unica = extrair_valor_regex(r"Demanda kW\s+([\d\.,]+)", texto)
        if dc_unica > 0: dados['demanda_contratada_fponta'] = dc_unica
        dados['demanda_contratada_ponta'] = 0.0
        
    if "B3" in dados['classificacao']:
        dados['demanda_contratada_ponta'] = 0.0
        dados['demanda_contratada_fponta'] = 0.0

    m_linha_principal = re.search(r"([A-Z]{3}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+R\$\s*([\d\.]+,\d{2})", texto)
    if m_linha_principal:
        dados['mes_referencia'] = m_linha_principal.group(1)
        dados['data_vencimento'] = m_linha_principal.group(2)
        dados['valor_total_fatura'] = limpar_numero(m_linha_principal.group(3))
    else:
        dados['mes_referencia'] = extrair_texto_regex(r"([A-Z]{3}/\d{4})\s+\d{2}/\d{2}/\d{4}\s+R\$", texto)
        dados['data_vencimento'] = extrair_texto_regex(r"[A-Z]{3}/\d{4}\s+(\d{2}/\d{2}/\d{4})\s+R\$", texto)
        dados['valor_total_fatura'] = 0.0

    dados['data_proxima_leitura'] = extrair_texto_regex(r"Próxima [Ll]eitura\s+(\d{2}/\d{2}/\d{4})", texto, padrao_falha="")
    
    leitura = re.search(r"\b\d{5,12}\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d{2,3}\b", texto)
    if leitura:
        dados['periodo_leitura_fim'] = leitura.group(1)
        dados['periodo_leitura_inicio'] = leitura.group(2)
    else:
        leitura_alt = re.search(r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d{2,3}", texto)
        dados['periodo_leitura_fim'] = leitura_alt.group(1) if leitura_alt else ""
        dados['periodo_leitura_inicio'] = leitura_alt.group(2) if leitura_alt else ""

    m_tusd_p = re.search(r"Consumo Ponta \[KWh\]\s*-\s*TUSD.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_tusd_p: dados['consumo_ponta'], dados['tarifa_aneel_cons_ponta_tusd'], dados['tarifa_trib_cons_ponta_tusd'], dados['valor_cons_ponta_tusd'] = [limpar_numero(x) for x in m_tusd_p.groups()]

    m_tusd_fp = re.search(r"Consumo Fora Ponta \[KWh\]\s*-\s*TUSD.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_tusd_fp: dados['consumo_fora_ponta'], dados['tarifa_aneel_cons_fponta_tusd'], dados['tarifa_trib_cons_fponta_tusd'], dados['valor_cons_fponta_tusd'] = [limpar_numero(x) for x in m_tusd_fp.groups()]

    m_te_p = re.search(r"Cons(?:umo)? Ponta\s*-\s*TE.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_te_p:
        vals = [limpar_numero(x) for x in m_te_p.groups()]
        dados['tarifa_aneel_cons_ponta_te'], dados['tarifa_trib_cons_ponta_te'], dados['valor_cons_ponta_te'] = vals[1], vals[2], vals[3]

    m_te_fp = re.search(r"Cons(?:umo)? F(?:ora)?\s*Ponta\s*TE.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_te_fp:
        vals = [limpar_numero(x) for x in m_te_fp.groups()]
        dados['tarifa_aneel_cons_fponta_te'], dados['tarifa_trib_cons_fponta_te'], dados['valor_cons_fponta_te'] = vals[1], vals[2], vals[3]

    if dados['consumo_fora_ponta'] == 0.0:
        m_tusd_b3 = re.search(r"Consumo Uso Sistema \[KWh\]\s*-\s*TUSD.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
        if m_tusd_b3: dados['consumo_fora_ponta'], dados['tarifa_aneel_cons_fponta_tusd'], dados['tarifa_trib_cons_fponta_tusd'], dados['valor_cons_fponta_tusd'] = [limpar_numero(x) for x in m_tusd_b3.groups()]

    if dados['valor_cons_fponta_te'] == 0.0:
        m_te_b3 = re.search(r"Consumo\s*-\s*TE.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
        if m_te_b3:
            vals = [limpar_numero(x) for x in m_te_b3.groups()]
            dados['tarifa_aneel_cons_fponta_te'], dados['tarifa_trib_cons_fponta_te'], dados['valor_cons_fponta_te'] = vals[1], vals[2], vals[3]

    dados['tipo_bandeira'] = extrair_texto_regex(r"Adicional Band (Verde|Amarela|Vermelha I|Vermelha II|Escassez Hídrica)", texto, padrao_falha="VERDE").upper()
    vp = extrair_valor_regex(r"Adicional Band.*?Ponta.*?kWh\s+([\d\.]+,\d{2})", texto)
    vfp = extrair_valor_regex(r"Adicional Band.*?FPonta.*?kWh\s+([\d\.]+,\d{2})", texto)
    dados['adicional_bandeira'] = vp + vfp

    linhas_ponta = re.findall(r"Demanda Ponta \[kW\]\s*-\s*TUSD.*?kW\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if len(linhas_ponta) >= 2:
        parsed = [[limpar_numero(x) for x in linha] for linha in linhas_ponta]
        parsed.sort(key=lambda x: x[2]) 
        dados['demanda_isenta_ponta'], dados['tarifa_aneel_dem_isenta_ponta'], dados['tarifa_trib_dem_isenta_ponta'], dados['valor_dem_isenta_ponta'] = parsed[0]
        dados['demanda_registrada_ponta'], dados['tarifa_aneel_dem_ponta'], dados['tarifa_trib_dem_ponta'], dados['valor_dem_ponta'] = parsed[1]
    elif len(linhas_ponta) == 1:
        dados['demanda_registrada_ponta'], dados['tarifa_aneel_dem_ponta'], dados['tarifa_trib_dem_ponta'], dados['valor_dem_ponta'] = [limpar_numero(x) for x in linhas_ponta[0]]

    linhas_fponta = re.findall(r"Demanda(?: F(?:ora)? Ponta)? \[kW\]\s*-\s*TUSD.*?kW\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if len(linhas_fponta) >= 2:
        parsed = [[limpar_numero(x) for x in linha] for linha in linhas_fponta]
        parsed.sort(key=lambda x: x[2])
        dados['demanda_isenta_fora_ponta'], dados['tarifa_aneel_dem_isenta_fponta'], dados['tarifa_trib_dem_isenta_fponta'], dados['valor_dem_isenta_fponta'] = parsed[0]
        dados['demanda_registrada_fora_ponta'], dados['tarifa_aneel_dem_fponta'], dados['tarifa_trib_dem_fponta'], dados['valor_dem_fponta'] = parsed[1]
    elif len(linhas_fponta) == 1:
        dados['demanda_registrada_fora_ponta'], dados['tarifa_aneel_dem_fponta'], dados['tarifa_trib_dem_fponta'], dados['valor_dem_fponta'] = [limpar_numero(x) for x in linhas_fponta[0]]

    m_reat_p = re.search(r"Consumo Reativo Exc Ponta.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_reat_p: dados['consumo_reativo_ponta'], dados['tarifa_aneel_cons_reativo_ponta'], dados['tarifa_trib_cons_reativo_ponta'], dados['valor_cons_reativo_ponta'] = [limpar_numero(x) for x in m_reat_p.groups()]

    m_reat_fp = re.search(r"Consumo Reativo Exc Fora Ponta.*?kWh\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_reat_fp: dados['consumo_reativo_fora_ponta'], dados['tarifa_aneel_cons_reativo_fponta'], dados['tarifa_trib_cons_reativo_fponta'], dados['valor_cons_reativo_fponta'] = [limpar_numero(x) for x in m_reat_fp.groups()]

    m_ultrap_p = re.search(r"Demanda Ultrap Ponta.*?kW\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_ultrap_p: dados['demanda_ultrapassagem_ponta'], dados['tarifa_aneel_dem_ultrap_ponta'], dados['tarifa_trib_dem_ultrap_ponta'], dados['valor_dem_ultrap_ponta'] = [limpar_numero(x) for x in m_ultrap_p.groups()]

    m_ultrap_fp = re.search(r"Demanda Ultrap(?:assagem)?(?: Fponta)?.*?kW\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_ultrap_fp: dados['demanda_ultrapassagem_fora_ponta'], dados['tarifa_aneel_dem_ultrap_fponta'], dados['tarifa_trib_dem_ultrap_fponta'], dados['valor_dem_ultrap_fponta'] = [limpar_numero(x) for x in m_ultrap_fp.groups()]

    m_dem_reat_p = re.search(r"Dem Reat Exc(?:ed)? Ponta.*?kW\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_dem_reat_p: dados['demanda_reativa_ponta'], dados['tarifa_aneel_dem_reativa_ponta'], dados['tarifa_trib_dem_reativa_ponta'], dados['valor_dem_reativa_ponta'] = [limpar_numero(x) for x in m_dem_reat_p.groups()]

    m_dem_reat_fp = re.search(r"Dem Reat Exc(?:ed)?(?: F(?:ora)? Ponta)?.*?kW\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)\s+([\d\.]+,\d+)", texto, re.IGNORECASE)
    if m_dem_reat_fp: dados['demanda_reativa_fora_ponta'], dados['tarifa_aneel_dem_reativa_fponta'], tarifa_trib_dem_reativa_fponta, dados['valor_dem_reativa_fponta'] = [limpar_numero(x) for x in m_dem_reat_fp.groups()]

    dados['cip'] = extrair_valor_regex(r"Contribuição Custeio IP-CIP.*?([\d\.]+,\d+)", texto)
    val_subtotal = extrair_valor_regex(r"Subtotal\s*([\d.,]+)", texto)
    if val_subtotal == 0.0:
        val_subtotal = extrair_valor_regex(r"Total Distribuidora\s*([\d.,]+)", texto)  
    if val_subtotal == 0.0:
        val_subtotal = extrair_valor_regex(r"ICMS\s+([\d.,]+)\s+\d{2}[.,]\d{2}", texto)
    dados['subtotal_fatura'] = val_subtotal
    dados['retencao_consumo_irrf'] = extrair_valor_regex(r"Retencao Consumo IRRF-.*?([\d\.]+,[\d]{2})-", texto)
    dados['retencao_demanda_irrf'] = extrair_valor_regex(r"Retencao Demanda IRRF-.*?([\d\.]+,[\d]{2})-", texto)
    dados['valor_total_pis'] = extrair_valor_regex(r"PIS/PASEP.*?\s([\d\.]+,\d+)$", texto)
    dados['valor_total_cofins'] = extrair_valor_regex(r"COFINS.*?\s([\d\.]+,\d+)$", texto)
    dados['valor_total_icms'] = extrair_valor_regex(r"ICMS.*?\s([\d\.]+,\d+)$", texto)
    
    valor_pagar_fim = extrair_valor_regex(r"Total a Pagar\s+([\d\.]+,\d{2})", texto)
    if valor_pagar_fim > 0:
        dados['valor_total_fatura'] = valor_pagar_fim
        
    return dados

def processar_pdf_cemig(arquivo_pdf):
    """Extrai os dados específicos da fatura da Comercializadora (CEMIG) no ACL."""
    with pdfplumber.open(arquivo_pdf) as pdf:
        texto = ""
        for i in range(min(2, len(pdf.pages))):
            texto_pagina = pdf.pages[i].extract_text()
            if texto_pagina:
                texto += texto_pagina + "\n"
        
    dados_cemig = {}
    
    # 1. Identificação Básica no PDF da CEMIG (Busca Flexível)
    match_uc = re.search(r"UNIDADE CONSUMIDORA[\s\S]*?(\d{10})", texto, re.IGNORECASE)
    uc_cemig_lida = match_uc.group(1).strip() if match_uc else ""
    
    match_mes = re.search(r"Referente a[\s\S]*?([A-Z]{3}/\d{4})", texto, re.IGNORECASE)
    dados_cemig['mes_referencia'] = match_mes.group(1).strip().upper() if match_mes else ""
    
    vencimento_bruto = extrair_texto_regex(r"Vencimento\s*(\d{2}/\d{2}/\d{4})", texto)
    dados_cemig['data_vencimento_acl'] = vencimento_bruto if vencimento_bruto else extrair_texto_regex(r"(\d{2}/\d{2}/\d{4})", texto) 
    
    # 2. BUSCA BLINDADA DA UC NO BANCO DE DADOS
    conexao_pdf = obter_conexao()
    c_pdf = conexao_pdf.cursor()
    
    if uc_cemig_lida:
        # Usar o LIKE '%numero%' garante que o sistema acha a UC mesmo que tenha '.0' ou espaços escondidos
        c_pdf.execute("""
            SELECT unidade_consumidora, nome_unidade, atividade 
            FROM cadastro_uc 
            WHERE uc_cemig LIKE %s
        """, (f"%{uc_cemig_lida}%",))
        res_uc = c_pdf.fetchone()
    else:
        res_uc = None
        
    conexao_pdf.close()
    
    if res_uc:
        dados_cemig['unidade_consumidora'] = res_uc[0] # Associa magicamente à UC Master da CPFL!
        dados_cemig['nome_unidade'] = res_uc[1]
        dados_cemig['atividade'] = res_uc[2]
        dados_cemig['uc_original'] = uc_cemig_lida
    else:
        # ALERTA VISUAL: Se não achar, não falha em silêncio. Avisa o utilizador!
        dados_cemig['unidade_consumidora'] = uc_cemig_lida
        dados_cemig['nome_unidade'] = "⚠️ VINCULAR UC CEMIG NO CADASTRO"
        dados_cemig['atividade'] = "Administrativa" 
        dados_cemig['uc_original'] = uc_cemig_lida
    
    # 3. Extração de Valores ACL
    linha_energia = re.search(r"Energia Ativa HFP.*?(?:kWh)\s+([\d\.]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if linha_energia:
        dados_cemig['consumo_energia_acl_kwh'] = limpar_numero(linha_energia.group(1))
        dados_cemig['tarifa_energia_acl'] = float(linha_energia.group(2).replace(',', '.'))
        dados_cemig['valor_energia_acl'] = round(dados_cemig['consumo_energia_acl_kwh'] * dados_cemig['tarifa_energia_acl'], 2)
    else:
        dados_cemig['consumo_energia_acl_kwh'] = 0.0
        dados_cemig['tarifa_energia_acl'] = 0.0
        dados_cemig['valor_energia_acl'] = 0.0

    # Valor Total a Pagar da Fatura CEMIG
    match_total = re.search(r"Total a pagar[\s\S]*?R\$\s*([\d\.,]+)", texto, re.IGNORECASE)
    
    if not match_total:
        # Plano B: Busca o valor no cabeçalho do PDF
        match_total = re.search(r"Valor a pagar \(R\$\)[\s\S]*?\d{2}/\d{2}/\d{4}\s+([\d\.,]+)", texto, re.IGNORECASE)
        
    if not match_total:
        # Plano C: Busca na linha de totalização dos itens
        match_total = re.search(r"TOTAL\s+([\d\.,]+)", texto, re.IGNORECASE)
        
    dados_cemig['valor_total_acl'] = limpar_numero(match_total.group(1)) if match_total else 0.0
    
    return dados_cemig

def processar_pdf_cpfl_acl(arquivo_pdf):
    with pdfplumber.open(arquivo_pdf) as pdf:
        texto = ""
        # LÊ TODAS AS PÁGINAS DO PDF (Garante a captura da Bandeira nas páginas finais)
        for page in pdf.pages:
            texto_pagina = page.extract_text()
            if texto_pagina:
                texto += texto_pagina + "\n"
            
    # Inicializa todas as colunas numéricas com 0.0
    chaves_numericas = [
        'demanda_contratada_ponta', 'demanda_contratada_fponta',
        'consumo_ponta', 'tarifa_aneel_cons_ponta_tusd', 'tarifa_trib_cons_ponta_tusd', 'valor_cons_ponta_tusd',
        'consumo_fora_ponta', 'tarifa_aneel_cons_fponta_tusd', 'tarifa_trib_cons_fponta_tusd', 'valor_cons_fponta_tusd',
        'tarifa_aneel_cons_ponta_te', 'tarifa_trib_cons_ponta_te', 'valor_cons_ponta_te',
        'tarifa_aneel_cons_fponta_te', 'tarifa_trib_cons_fponta_te', 'valor_cons_fponta_te',
        'demanda_registrada_ponta', 'tarifa_aneel_dem_ponta', 'tarifa_trib_dem_ponta', 'valor_dem_ponta',
        'demanda_isenta_ponta', 'tarifa_aneel_dem_isenta_ponta', 'tarifa_trib_dem_isenta_ponta', 'valor_dem_isenta_ponta',
        'demanda_registrada_fora_ponta', 'tarifa_aneel_dem_fponta', 'tarifa_trib_dem_fponta', 'valor_dem_fponta',
        'demanda_isenta_fora_ponta', 'tarifa_aneel_dem_isenta_fponta', 'tarifa_trib_dem_isenta_fponta', 'valor_dem_isenta_fponta',
        'consumo_reativo_ponta', 'tarifa_aneel_cons_reativo_ponta', 'tarifa_trib_cons_reativo_ponta', 'valor_cons_reativo_ponta',
        'consumo_reativo_fora_ponta', 'tarifa_aneel_cons_reativo_fponta', 'tarifa_trib_cons_reativo_fponta', 'valor_cons_reativo_fponta',
        'demanda_ultrapassagem_ponta', 'tarifa_aneel_dem_ultrap_ponta', 'tarifa_trib_dem_ultrap_ponta', 'valor_dem_ultrap_ponta',
        'demanda_ultrapassagem_fora_ponta', 'tarifa_aneel_dem_ultrap_fponta', 'tarifa_trib_dem_ultrap_fponta', 'valor_dem_ultrap_fponta',
        'demanda_reativa_ponta', 'tarifa_aneel_dem_reativa_ponta', 'tarifa_trib_dem_reativa_ponta', 'valor_dem_reativa_ponta',
        'demanda_reativa_fora_ponta', 'tarifa_aneel_dem_reativa_fponta', 'tarifa_trib_dem_reativa_fponta', 'valor_dem_reativa_fponta',
        'subtotal_fatura', 'cip', 'retencao_consumo_irrf', 'retencao_demanda_irrf', 'valor_total_pis', 'valor_total_cofins', 'valor_total_icms',
        'consumo_energia_acl_kwh', 'tarifa_energia_acl', 'valor_total_acl'
    ]
    dados = {k: 0.0 for k in chaves_numericas}
    
    dados['periodo_leitura_inicio'] = ""
    dados['periodo_leitura_fim'] = ""
    dados['data_proxima_leitura'] = ""
    dados['tipo_bandeira'] = "VERDE"
    dados['adicional_bandeira'] = 0.0
    dados['data_vencimento_acl'] = ""
    
    # 2. Classificação (Verde Livre ou Azul Livre)
    classif_match = re.search(r"Classificação(?:[:\.])?\s*(.*?)(?:\n|Serviço|Autarquia)", texto, re.IGNORECASE)
    classificacao_bruta = classif_match.group(1).strip().upper() if classif_match else ""
    
    if "VERDE" in classificacao_bruta and "LIVRE" in classificacao_bruta:
        dados['classificacao'] = "Tarifa Verde Livre-A4"
    elif "AZUL" in classificacao_bruta and "LIVRE" in classificacao_bruta:
        dados['classificacao'] = "Tarifa Azul Livre-A4"
    elif "LIVRE" in classificacao_bruta:
        if re.search(r"Uso Sist Distr Ponta", texto, re.IGNORECASE):
            dados['classificacao'] = "Tarifa Azul Livre-A4"
        else:
            dados['classificacao'] = "Tarifa Verde Livre-A4"
    else:
        dados['classificacao'] = "Tarifa Verde Livre-A4"

    # 3. Unidade Consumidora (UC Nova CPFL)
    uc_match = re.search(r"DEPARTAMENTO DE AGUA E ESGOTO DAE\s*(\d{3}\.\d{3}\.\d{3}-\d{2})", texto)
    if not uc_match:
        uc_match = re.search(r"(\d{3}\.\d{3}\.\d{3}-\d{2})", texto)
    dados['unidade_consumidora'] = uc_match.group(1).strip() if uc_match else ""
    
    # 4. Datas
    m_venc_ref = re.search(r"([A-Z]{3}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+R\$", texto)
    if m_venc_ref:
        dados['mes_referencia'] = m_venc_ref.group(1)
        dados['data_vencimento'] = m_venc_ref.group(2)
    else:
        dados['data_vencimento'] = extrair_texto_regex(r"Data de Vencimento\s*(\d{2}/\d{2}/\d{4})", texto)
        dados['mes_referencia'] = extrair_texto_regex(r"Referente a\s*([A-Z]{3}/\d{4})", texto)

    leitura_ant = re.search(r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+\d{2,3}", texto)
    if leitura_ant:
        dados['periodo_leitura_inicio'] = leitura_ant.group(1)
        dados['periodo_leitura_fim'] = leitura_ant.group(2)
        
    dados['data_proxima_leitura'] = extrair_texto_regex(r"Próxima Leitura\s*(\d{2}/\d{2}/\d{4})", texto)

    # 5. Busca Contratos
    conexao_pdf = obter_conexao()
    c_pdf = conexao_pdf.cursor()
    c_pdf.execute("SELECT nome_unidade, atividade, demanda_contratada_ponta, demanda_contratada_fponta FROM cadastro_uc WHERE unidade_consumidora = %s", (dados['unidade_consumidora'],))
    res_uc = c_pdf.fetchone()
    conexao_pdf.close()
    
    dados['nome_unidade'] = res_uc[0] if res_uc else "Não Cadastrada"
    dados['atividade'] = res_uc[1] if res_uc else "Administrativa" 
    dados['demanda_contratada_ponta'] = res_uc[2] if res_uc else 0.0
    dados['demanda_contratada_fponta'] = res_uc[3] if res_uc else 0.0

    # 6. Extração de Itens Faturados (TUSD, Demandas e Reativos)
    # Consumo Ponta
    m_tusd_p = re.search(r"Tusd Enc Cons Ponta.*?kWh\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if m_tusd_p: 
        dados['consumo_ponta'], dados['tarifa_aneel_cons_ponta_tusd'], dados['tarifa_trib_cons_ponta_tusd'], dados['valor_cons_ponta_tusd'] = [limpar_numero(x) for x in m_tusd_p.groups()]

    # Consumo Fora Ponta
    m_tusd_fp = re.search(r"Tusd Enc Cons F Ponta.*?kWh\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if m_tusd_fp: 
        dados['consumo_fora_ponta'], dados['tarifa_aneel_cons_fponta_tusd'], dados['tarifa_trib_cons_fponta_tusd'], dados['valor_cons_fponta_tusd'] = [limpar_numero(x) for x in m_tusd_fp.groups()]

    # Demanda Ponta
    linhas_ponta = re.findall(r"Uso Sist Distr Ponta.*?kW\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if len(linhas_ponta) >= 2:
        parsed = [[limpar_numero(x) for x in linha] for linha in linhas_ponta]
        parsed.sort(key=lambda x: x[0], reverse=True)
        dados['demanda_registrada_ponta'], dados['tarifa_aneel_dem_ponta'], dados['tarifa_trib_dem_ponta'], dados['valor_dem_ponta'] = parsed[0]
        dados['demanda_isenta_ponta'], dados['tarifa_aneel_dem_isenta_ponta'], dados['tarifa_trib_dem_isenta_ponta'], dados['valor_dem_isenta_ponta'] = parsed[1]
    elif len(linhas_ponta) == 1:
        dados['demanda_registrada_ponta'], dados['tarifa_aneel_dem_ponta'], dados['tarifa_trib_dem_ponta'], dados['valor_dem_ponta'] = [limpar_numero(x) for x in linhas_ponta[0]]

    # Demanda Fora Ponta
    linhas_fponta = re.findall(r"Uso Sist Distr F Ponta.*?kW\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if len(linhas_fponta) >= 2:
        parsed = [[limpar_numero(x) for x in linha] for linha in linhas_fponta]
        parsed.sort(key=lambda x: x[0], reverse=True)
        dados['demanda_registrada_fora_ponta'], dados['tarifa_aneel_dem_fponta'], dados['tarifa_trib_dem_fponta'], dados['valor_dem_fponta'] = parsed[0]
        dados['demanda_isenta_fora_ponta'], dados['tarifa_aneel_dem_isenta_fponta'], dados['tarifa_trib_dem_isenta_fponta'], dados['valor_dem_isenta_fponta'] = parsed[1]
    elif len(linhas_fponta) == 1:
        dados['demanda_registrada_fora_ponta'], dados['tarifa_aneel_dem_fponta'], dados['tarifa_trib_dem_fponta'], dados['valor_dem_fponta'] = [limpar_numero(x) for x in linhas_fponta[0]]

    # Reativos
    m_reat_p = re.search(r"USD Consumo Reativo Ponta.*?kWh\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if m_reat_p: dados['consumo_reativo_ponta'], dados['tarifa_aneel_cons_reativo_ponta'], dados['tarifa_trib_cons_reativo_ponta'], dados['valor_cons_reativo_ponta'] = [limpar_numero(x) for x in m_reat_p.groups()]

    m_reat_fp = re.search(r"USD Consumo Reativo Fora Ponta.*?kWh\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if m_reat_fp: dados['consumo_reativo_fora_ponta'], dados['tarifa_aneel_cons_reativo_fponta'], dados['tarifa_trib_cons_reativo_fponta'], dados['valor_cons_reativo_fponta'] = [limpar_numero(x) for x in m_reat_fp.groups()]

    m_dem_reat_p = re.search(r"Dem Reat Exc Ponta.*?kW\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if m_dem_reat_p: dados['demanda_reativa_ponta'], dados['tarifa_aneel_dem_reativa_ponta'], dados['tarifa_trib_dem_reativa_ponta'], dados['valor_dem_reativa_ponta'] = [limpar_numero(x) for x in m_dem_reat_p.groups()]

    m_dem_reat_fp = re.search(r"Dem Reat Exc FPonta.*?kW\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)", texto, re.IGNORECASE)
    if m_dem_reat_fp: dados['demanda_reativa_fora_ponta'], dados['tarifa_aneel_dem_reativa_fponta'], dados['tarifa_trib_dem_reativa_fponta'], dados['valor_dem_reativa_fponta'] = [limpar_numero(x) for x in m_dem_reat_fp.groups()]

    # 7. EXTRAÇÃO ATUALIZADA DA BANDEIRA TARIFÁRIA
    match_bandeira = re.search(
        r"Energia Ativa[^\n]*?\b(Verde|Amarela|Vermelha\s*I{1,2}|Vermelha|Escassez Hídrica)\b",
        texto,
        re.IGNORECASE
    )
    if not match_bandeira:
        match_bandeira = re.search(
            r"\b(Verde|Amarela|Vermelha\s*I{1,2}|Vermelha|Escassez Hídrica)\b\s*\d{1,3}\s*Dias",
            texto,
            re.IGNORECASE
        )
    
    if match_bandeira:
        dados['tipo_bandeira'] = match_bandeira.group(1).strip().upper()
    else:
        dados['tipo_bandeira'] = "VERDE"

    vp = extrair_valor_regex(r"CDE Escassez Hídrica Ponta.*?kWh\s+[\d\.,]+\s+[\d\.,]+\s+[\d\.,]+\s+([\d\.,]+)", texto)
    vfp = extrair_valor_regex(r"CDE Escassez Hídrica F(?:ora)? Ponta.*?kWh\s+[\d\.,]+\s+[\d\.,]+\s+[\d\.,]+\s+([\d\.,]+)", texto)

    # 8. Impostos e Totais
    dados['cip'] = extrair_valor_regex(r"Contribuição Custeio IP-CIP.*?\s([\d\.,]+)", texto)
    
    val_subtotal = extrair_valor_regex(r"Total Distribuidora\s*([\d\.,]+)", texto)
    if val_subtotal == 0.0: val_subtotal = extrair_valor_regex(r"Subtotal\s*([\d\.,]+)", texto)  
    dados['subtotal_fatura'] = val_subtotal
    
    dados['retencao_consumo_irrf'] = extrair_valor_regex(r"Retencao Consumo IRRF.*?\s([\d\.,]+)-", texto)
    dados['retencao_demanda_irrf'] = extrair_valor_regex(r"Retencao Demanda IRRF.*?\s([\d\.,]+)-", texto)
    dados['valor_total_pis'] = extrair_valor_regex(r"PIS/PASEP.*?\s([\d\.,]+)$", texto)
    dados['valor_total_cofins'] = extrair_valor_regex(r"COFINS.*?\s([\d\.,]+)$", texto)
    dados['valor_total_icms'] = extrair_valor_regex(r"ICMS.*?\s([\d\.,]+)$", texto)
    
    valor_pagar_fim = extrair_valor_regex(r"Total a Pagar\s*([\d\.,]+)", texto)
    if valor_pagar_fim > 0: dados['valor_total_fatura'] = valor_pagar_fim
        
    return dados
    
# --- 4. INTERFACE ---
aba_dash, aba_controle, aba_dados, aba_espelho, aba_pdf, aba_config = st.tabs(["📈 Dashboard", "💰 Controle Financeiro", "📊 Banco de Dados", "📑 Espelho de Fatura", "📄 Upload de Fatura", "⚙️ Configurações"])

# ==========================================
# ABA DASHBOARD
# ==========================================
with aba_dash:
    df_dash = carregar_dados()
    
    if not df_dash.empty:
        st.markdown("##### ⚡ Business Intelligence - Consumo DAE")
        st.markdown("💡 **Dica:** Segure a tecla **SHIFT** para seleção múltipla nos gráficos. Utilize os filtros dinâmicos dos gráficos.")
        
        # 1. Preparação dos Dados
        df_dash['Ano'] = df_dash['Data Referência Oculta'].dt.year.astype(str)
        df_dash['Mes_Nome'] = df_dash['Data Referência Oculta'].dt.strftime('%B')
        df_dash['Mes_Num'] = df_dash['Data Referência Oculta'].dt.month
        
        # 2. Inicialização da Memória de Cliques
        if 'clique_ano' not in st.session_state: st.session_state.clique_ano = []
        if 'clique_mes' not in st.session_state: st.session_state.clique_mes = []
        if 'clique_uc' not in st.session_state: st.session_state.clique_uc = []

        st.divider()

        # --- SELETORES SUPERIORES E BOTÃO DE RESET ---
        dic_parametros = {
            "Consumo Total (kWh)": "Total Consumo",
            "Valor Total Fatura (R$)": "Valor Total Fatura",
            "Valor Total Consumo (R$)": "Valor Total Consumo",
            "Valor Total Demanda (R$)": "Valor Total Dem.",
            "Valor Total Desv. Dem. (R$)": "Valor Total Desv. Dem.",
            "Valor Total Reativo (R$)": "Valor Total Reativo"
        }
        
        # Ajuste de layout: Indicador (2.5), Classificação (2.5), Espaço (4), Botão (1.5)
        col_ind, col_cla, col_vazio, col_btn = st.columns([2.5, 2.5, 4, 1.5]) 
        
        param_nome = col_ind.selectbox("🎯 **Indicador:**", list(dic_parametros.keys()))
        param_coluna = dic_parametros[param_nome]
        is_dinheiro = "(R$)" in param_nome

        # Ajuste: Mudança para multiselect
        opcoes_classes = sorted(list(df_dash['Classificação'].unique()))
        filtro_classe_fixo = col_cla.multiselect("🏷️ **Classificação:**", options=opcoes_classes, placeholder="Todas as classes")

        # Botão de Reset (Extrema direita)
        if st.session_state.clique_ano or st.session_state.clique_mes or st.session_state.clique_uc:
            col_btn.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if col_btn.button("🧹 Limpar Filtros", use_container_width=True):
                st.session_state.clique_ano = []; st.session_state.clique_mes = []; st.session_state.clique_uc = []
                st.rerun()

        # 3. Lógica de Filtragem (Cross-Filtering Completo)
        df_base_dash = df_dash.copy()
        
        # Filtro Fixo (Classificação)
        if filtro_classe_fixo:
            df_base_dash = df_base_dash[df_base_dash['Classificação'].isin(filtro_classe_fixo)]
            
        # --- PREPARAÇÃO DOS DATAFRAMES INDIVIDUAIS ---
        # DF do Gráfico 1 (Ano): Ignora clique_ano, aplica mês e UC
        df_para_ano = df_base_dash.copy()
        if st.session_state.clique_mes:
            df_para_ano = df_para_ano[df_para_ano['Mes_Nome'].isin(st.session_state.clique_mes)]
        if st.session_state.clique_uc:
            df_para_ano = df_para_ano[df_para_ano['Nome da Unidade'].isin(st.session_state.clique_uc)]

        # DF do Gráfico 2 (Mês): Ignora clique_mes, aplica ano e UC
        df_para_mes = df_base_dash.copy()
        if st.session_state.clique_ano:
            df_para_mes = df_para_mes[df_para_mes['Ano'].isin(st.session_state.clique_ano)]
        if st.session_state.clique_uc:
            df_para_mes = df_para_mes[df_para_mes['Nome da Unidade'].isin(st.session_state.clique_uc)]

        # DF do Gráfico 3 (UC): Ignora clique_uc, aplica ano e mês
        df_para_uc = df_base_dash.copy()
        if st.session_state.clique_ano:
            df_para_uc = df_para_uc[df_para_uc['Ano'].isin(st.session_state.clique_ano)]
        if st.session_state.clique_mes:
            df_para_uc = df_para_uc[df_para_uc['Mes_Nome'].isin(st.session_state.clique_mes)]

        st.write("")
        col_graf1, col_graf2 = st.columns(2)
        
        # --- GRÁFICO 1: Análise Anual ---
        df_ano = df_para_ano.groupby('Ano')[param_coluna].sum().reset_index()
        
        # Gera a lista de cores dinamicamente (List Comprehension)
        cores_ano = ["#0055A5" if not st.session_state.clique_ano or ano in st.session_state.clique_ano else "#D3D3D3" for ano in df_ano['Ano']]
                
        fig_ano = px.bar(df_ano, x='Ano', y=param_coluna, text_auto='.3s', title=f"{param_nome} por Ano")
        fig_ano.update_layout(xaxis_title=None, yaxis_title=param_nome, xaxis={'type': 'category'})
        
        if is_dinheiro:
            fig_ano.update_traces(marker_color=cores_ano, texttemplate='R$ %{y:.3s}', textposition='outside')
        else:
            fig_ano.update_traces(marker_color=cores_ano, texttemplate='%{y:.3s}', textposition='outside')

        evento_ano = col_graf1.plotly_chart(fig_ano, use_container_width=True, on_select="rerun", selection_mode=("points", "box", "lasso"))
        
        if evento_ano and len(evento_ano.selection.get("points", [])) > 0:
            anos_sel = sorted(list(set([str(pt["x"]) for pt in evento_ano.selection["points"]])))
            anos_atuais = sorted(st.session_state.clique_ano)
            
            # Se clicou na mesma barra que já estava filtrada, ele "desliga" o filtro
            if anos_sel == anos_atuais:
                st.session_state.clique_ano = []
                st.rerun()
            # Se clicou em uma barra nova, ele aplica o filtro novo
            elif anos_sel != anos_atuais:
                st.session_state.clique_ano = anos_sel
                st.rerun()

        # --- GRÁFICO 2: Sazonalidade Mensal ---
        df_mes_ciclo = df_para_mes.groupby(['Mes_Num', 'Mes_Nome'])[param_coluna].sum().reset_index().sort_values('Mes_Num')
        
        fig_mes = px.line(df_mes_ciclo, x='Mes_Nome', y=param_coluna, markers=True, title=f"Sazonalidade Mensal ({param_nome})")
        fig_mes.update_layout(xaxis_title=None, yaxis_title=param_nome)
        fig_mes.update_yaxes(rangemode="tozero")
        
        # Cores e tamanhos para os marcadores da linha
        cores_mes = ["#0055A5" if not st.session_state.clique_mes or mes in st.session_state.clique_mes else "#D3D3D3" for mes in df_mes_ciclo['Mes_Nome']]
        tamanho_mes = [10 if not st.session_state.clique_mes or mes in st.session_state.clique_mes else 6 for mes in df_mes_ciclo['Mes_Nome']]
        
        # Atualiza a linha e os pontos
        cor_linha = "#A0C4E1" if st.session_state.clique_mes else "#0055A5" # Deixa a linha mais clara se houver filtro
        fig_mes.update_traces(line=dict(color=cor_linha, width=3), marker=dict(color=cores_mes, size=tamanho_mes))
        
        if not df_mes_ciclo.empty:
            media_val = df_mes_ciclo[param_coluna].mean()
            fmt = f"Média: R$ {media_val:,.2f}" if is_dinheiro else f"Média: {media_val:,.0f}"
            fmt = fmt.replace(',', 'X').replace('.', ',').replace('X', '.')
            fig_mes.add_hline(y=media_val, line_dash="dash", line_color="#FF4B4B", annotation_text=fmt, annotation_position="top right")

        evento_mes = col_graf2.plotly_chart(fig_mes, use_container_width=True, on_select="rerun", selection_mode=("points", "box", "lasso"))
        
        if evento_mes and len(evento_mes.selection.get("points", [])) > 0:
            meses_sel = sorted(list(set([str(pt["x"]) for pt in evento_mes.selection["points"]])))
            meses_atuais = sorted(st.session_state.clique_mes)
            
            if meses_sel == meses_atuais:
                st.session_state.clique_mes = []
                st.rerun()
            elif meses_sel != meses_atuais:
                st.session_state.clique_mes = meses_sel
                st.rerun()

        st.divider()
        
        # --- GRÁFICO 3: Participação Total por Unidade ---     
        df_unidades = df_para_uc.groupby('Nome da Unidade')[param_coluna].sum().reset_index()
        total_indicador = df_unidades[param_coluna].sum()
        
        if total_indicador > 0:
            df_unidades['Percentual'] = (df_unidades[param_coluna] / total_indicador) * 100
        else:
            df_unidades['Percentual'] = 0
            
        df_unidades = df_unidades.sort_values(param_coluna, ascending=False).head(30)
        
        # Gera a lista de cores dinamicamente
        cores_uc = ["#0055A5" if not st.session_state.clique_uc or uc in st.session_state.clique_uc else "#D3D3D3" for uc in df_unidades['Nome da Unidade']]

        fig_unidades = px.bar(df_unidades, x='Nome da Unidade', y=param_coluna, title=f"📊 Top 30 Unidades por {param_nome}")
        
        fig_unidades.update_layout(
            xaxis_title=None, yaxis_title=param_nome, xaxis_tickangle=-45, 
            margin=dict(b=120), font=dict(size=14)
        )
        
        if is_dinheiro:
            fig_unidades.update_traces(marker_color=cores_uc, customdata=df_unidades['Percentual'], texttemplate='R$ %{y:.3s}<br>(%{customdata:.1f}%)', textposition='outside', textfont_size=13)
        else:
            fig_unidades.update_traces(marker_color=cores_uc, customdata=df_unidades['Percentual'], texttemplate='%{y:.3s}<br>(%{customdata:.1f}%)', textposition='outside', textfont_size=13)
        
        evento_uc = st.plotly_chart(fig_unidades, use_container_width=True, on_select="rerun", selection_mode=("points", "box", "lasso"))
        
        if evento_uc and len(evento_uc.selection.get("points", [])) > 0:
            ucs_sel = sorted(list(set([str(pt["x"]) for pt in evento_uc.selection["points"]])))
            ucs_atuais = sorted(st.session_state.clique_uc)
            
            if ucs_sel == ucs_atuais:
                st.session_state.clique_uc = []
                st.rerun()
            elif ucs_sel != ucs_atuais:
                st.session_state.clique_uc = ucs_sel
                st.rerun()

        # --- GRÁFICO 4: Análise de Demanda (Registrada vs Contratada) ---
        st.divider()
        st.markdown("##### 📉 Análise de Demanda: Registrada vs Contratada (Fora Ponta)")
        st.markdown("💡 **Dica:** Esta análise revela o comportamento real frente ao contrato. Selecione uma **Unidade Consumidora específica** no gráfico acima para visualizar o teto contratual individualizado.")

        # 1. Cria um DataFrame totalmente filtrado (obedecendo a TODOS os cliques ativos)
        df_demanda_filtrado = df_base_dash.copy()
        
        if st.session_state.clique_ano:
            df_demanda_filtrado = df_demanda_filtrado[df_demanda_filtrado['Ano'].isin(st.session_state.clique_ano)]
        if st.session_state.clique_mes:
            df_demanda_filtrado = df_demanda_filtrado[df_demanda_filtrado['Mes_Nome'].isin(st.session_state.clique_mes)]
        if st.session_state.clique_uc:
            df_demanda_filtrado = df_demanda_filtrado[df_demanda_filtrado['Nome da Unidade'].isin(st.session_state.clique_uc)]

        # 2. Prepara os dados agrupando cronologicamente
        df_demanda = df_demanda_filtrado.groupby('Data Referência Oculta').agg({
            'Dem. Reg. F.Ponta': 'sum',
            'Dem. Contr. F.Ponta': 'sum',
            'Mês Referência': 'first' # Mantém o rótulo do mês para o eixo X
        }).reset_index().sort_values('Data Referência Oculta')

        if not df_demanda.empty:
            # 3. Constrói o Gráfico Combinado usando Graph Objects (go)
            fig_demanda = go.Figure()

            # Adiciona as Barras (Demanda Registrada/Utilizada)
            fig_demanda.add_trace(
                go.Bar(
                    x=df_demanda['Mês Referência'], 
                    y=df_demanda['Dem. Reg. F.Ponta'],
                    name='Registrada (kW)', 
                    marker_color='#0055A5',
                    text=df_demanda['Dem. Reg. F.Ponta'],
                    textposition='auto',
                    texttemplate='%{text:.0f}'
                )
            )

            # Adiciona a Linha (Demanda Contratada/Limite)
            fig_demanda.add_trace(
                go.Scatter(
                    x=df_demanda['Mês Referência'], 
                    y=df_demanda['Dem. Contr. F.Ponta'],
                    mode='lines+markers', 
                    name='Contratada (kW)', 
                    line=dict(color='#FF4B4B', width=3, dash='dash'),
                    marker=dict(size=8, color='#FF4B4B')
                )
            )

            # 4. Estiliza o layout para ficar com aparência corporativa
            fig_demanda.update_layout(
                xaxis_title=None,
                yaxis_title="Demanda (kW)",
                barmode='group',
                hovermode="x unified",
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                )
            )

            # Exibe o gráfico na tela
            st.plotly_chart(fig_demanda, use_container_width=True)
        else:
            st.info("Não há dados de demanda para os filtros selecionados.")

# ==========================================
# ABA CONTROLE E AUDITORIA
# ==========================================
with aba_controle:
    st.markdown("##### 🔍 Painel de Controle Financeiro")
    
    # 1. Carregar dados básicos e cadastros ativos
    df_faturas = carregar_dados()
    conexao = obter_conexao()
    # Puxa a coluna uc_cemig para podermos identificar quem migrou para o ACL
    df_cadastro = pd.read_sql_query("SELECT unidade_consumidora, nome_unidade, status, dia_vencimento, uc_cemig FROM cadastro_uc WHERE status = 'ATIVA'", conexao)
    
    if df_faturas.empty:
        st.info("Nenhuma fatura carregada para auditoria.")
        conexao.close()
    else:
        # --- SEÇÃO 1: FILTRO E MÉTRICAS GERAIS ---
        col_filtro, _ = st.columns([1, 3]) 
        meses_disponiveis = df_faturas.sort_values('Data Referência Oculta', ascending=False)['Mês Referência'].unique().tolist()
        with col_filtro:
            mes_auditoria = st.selectbox("📅 Selecione o Mês:", meses_disponiveis)
        
        # Filtra todas as faturas dadas daquela competência de referência
        df_mes = df_faturas[df_faturas['Mês Referência'] == mes_auditoria]
        
        c1, c2, c3 = st.columns(3)
        c1.metric("Faturas no Mês", f"{len(df_mes)}")
        c2.metric("Total no Mês", f"R$ {df_mes['Valor Total Fatura'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        st.divider()

        # Criação das sub-abas internas (Sub-aba Fluxo de Vencimentos excluída com sucesso!)
        tab_relatorio, tab_pendencias = st.tabs(["📝 Gerar Relatório Financeiro", "🚨 Pendências de Carga"])

        # --- SUB-ABA 1: GERADOR DE RELATÓRIO ---
        with tab_relatorio:
            # Filtro de segurança: Isola estritamente as faturas da CPFL para o fechamento financeiro
            df_mes_cpfl = df_mes[df_mes['Classificação'] != 'Mercado Livre - ACL'].copy()
            
            # Busca faturas já enviadas para filtrar a visualização
            df_enviados = pd.read_sql_query(f"SELECT id, unidade_consumidora, data_envio, valor_fatura FROM historico_financeiro WHERE mes_referencia = '{mes_auditoria}'", conexao)
            ucs_enviadas = df_enviados['unidade_consumidora'].tolist()
            
            # Filtra apenas as faturas da distribuidora que ainda não foram enviadas
            df_pendente_envio = df_mes_cpfl[~df_mes_cpfl['UC'].isin(ucs_enviadas)].copy()
            
            if not df_pendente_envio.empty:
                # --- CÁLCULOS FINANCEIROS ---
                df_pendente_envio['Valor IRRF (-)'] = df_pendente_envio['Retenção Cons. IRRF'] + df_pendente_envio['Retenção Dem. IRRF']
                cols_energia = ['Valor Total Consumo', 'Valor Total Dem.', 'Valor Total Dem. Isenta', 'Valor Total Dem. Ultrap.', 'Valor Total Reativo', 'Adicional Bandeira']
                
                if 'Subtotal PDF' in df_pendente_envio.columns:
                    df_pendente_envio['Subtotal'] = df_pendente_envio.apply(lambda r: r['Subtotal PDF'] if r['Subtotal PDF'] > 0 else r[cols_energia].sum(), axis=1)
                else:
                    df_pendente_envio['Subtotal'] = df_pendente_envio[cols_energia].sum(axis=1)
                
                df_pendente_envio['Lançamentos Diversos'] = (df_pendente_envio['Valor Total Fatura'] - df_pendente_envio['Subtotal'] - df_pendente_envio['CIP'] + df_pendente_envio['Valor IRRF (-)']).round(2)
                df_pendente_envio['Lançamentos Diversos'] = df_pendente_envio['Lançamentos Diversos'].apply(lambda x: 0.0 if abs(x) <= 0.05 else x)
                
                # Mapeamento dinâmico inteligente para suportar tanto 'Vencimento' como 'Vencimento CPFL'
                col_venc_ativa = 'Vencimento CPFL' if 'Vencimento CPFL' in df_pendente_envio.columns else 'Vencimento'
                
                df_pendente_envio['Data_Ord'] = pd.to_datetime(df_pendente_envio[col_venc_ativa], format='%d/%m/%Y', errors='coerce')
                df_pendente_envio = df_pendente_envio.sort_values('Data_Ord')

                st.info(f"Existem **{len(df_pendente_envio)}** faturas da CPFL prontas para fechamento de lote.")
                colunas_fin = ['UC', 'Nome da Unidade', 'Mês Referência', col_venc_ativa, 'CIP', 'Subtotal', 'Valor IRRF (-)', 'Lançamentos Diversos', 'Valor Total Fatura']

                # --- EXIBIÇÃO NA TELA ---
                for atividade in sorted(df_pendente_envio['Atividade'].unique()):
                    with st.expander(f"🏢 SETOR: {atividade.upper()}", expanded=True):
                        df_ativ = df_pendente_envio[df_pendente_envio['Atividade'] == atividade].copy()
                        
                        st.markdown("##### 📝 Detalhamento de faturas")
                        df_detalhe = df_ativ[colunas_fin].copy()
                        for col in ['CIP', 'Subtotal', 'Valor IRRF (-)', 'Lançamentos Diversos', 'Valor Total Fatura']:
                            df_detalhe[col] = df_detalhe[col].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                        
                        st.dataframe(df_detalhe, hide_index=True, use_container_width=True)
                        
                        st.markdown("##### 📊 Resumo de pagamentos por data")
                        df_resumo = df_ativ.groupby(col_venc_ativa)['Valor Total Fatura'].sum().reset_index()
                        df_resumo['D_Ord'] = pd.to_datetime(df_resumo[col_venc_ativa], format='%d/%m/%Y', errors='coerce')
                        df_resumo = df_resumo.sort_values('D_Ord').drop(columns=['D_Ord'])
                        df_res_show = df_resumo.copy()
                        df_res_show.columns = ['Data de Vencimento', 'Valor Total']
                        df_res_show['Valor Total'] = df_res_show['Valor Total'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                        st.dataframe(df_res_show, hide_index=True, use_container_width=True)

                # --- GERAÇÃO DO EXCEL EM MEMÓRIA ---
                def finalizar_lote_db(dados, mes, col_venc):
                    conn = obter_conexao()
                    cursor = conn.cursor()
                    for _, row in dados.iterrows():
                        cursor.execute(
                            "INSERT INTO historico_financeiro (unidade_consumidora, mes_referencia, valor_fatura, vencimento) VALUES (%s, %s, %s, %s)",
                            (row['UC'], row['Mês Referência'], row['Valor Total Fatura'], row[col_venc])
                        )
                    conn.commit()
                    conn.close()

                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    pd.DataFrame().to_excel(writer, index=False, header=False, sheet_name='Relatorio_Financeiro')
                    ws = writer.sheets['Relatorio_Financeiro']
                    
                    header_fill = PatternFill(start_color="002060", fill_type="solid")
                    sector_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
                    font_white = Font(bold=True, color="FFFFFF")
                    font_bold = Font(bold=True)
                    center_align = Alignment(horizontal="center", vertical="center")
                    
                    row_idx = 1
                    
                    for atividade in sorted(df_pendente_envio['Atividade'].unique()):
                        df_ativ = df_pendente_envio[df_pendente_envio['Atividade'] == atividade].copy()
                        
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
                        c_setor = ws.cell(row=row_idx, column=1, value=f"SETOR: {atividade.upper()}")
                        c_setor.fill = sector_fill; c_setor.font = font_bold; c_setor.alignment = center_align
                        row_idx += 1
                        
                        for col_num, col_name in enumerate(colunas_fin, 1):
                            c_head = ws.cell(row=row_idx, column=col_num, value=col_name)
                            c_head.fill = header_fill; c_head.font = font_white; c_head.alignment = center_align
                        row_idx += 1
                        
                        for _, r in df_ativ.iterrows():
                            try:
                                uc_segura = int(float(r['UC']))
                            except (ValueError, TypeError):
                                uc_segura = str(r['UC']).strip()
                                
                            ws.cell(row=row_idx, column=1, value=uc_segura)
                            ws.cell(row=row_idx, column=2, value=str(r['Nome da Unidade']))
                            ws.cell(row=row_idx, column=3, value=str(r['Mês Referência']))
                            
                            try:
                                c_venc = ws.cell(row=row_idx, column=4, value=pd.to_datetime(r[col_venc_ativa], format='%d/%m/%Y'))
                                c_venc.number_format = 'DD/MM/YYYY'
                            except:
                                ws.cell(row=row_idx, column=4, value=str(r[col_venc_ativa]))
                            
                            for i, col_name in enumerate(['CIP', 'Subtotal', 'Valor IRRF (-)', 'Lançamentos Diversos', 'Valor Total Fatura'], 5):
                                c_val = ws.cell(row=row_idx, column=i, value=float(r[col_name]))
                                c_val.number_format = 'R$ #,##0.00'
                            row_idx += 1
                        
                        row_idx += 1
                        
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                        c_res = ws.cell(row=row_idx, column=1, value=f"RESUMO: {atividade.upper()}")
                        c_res.fill = sector_fill; c_res.font = font_bold; c_res.alignment = center_align
                        row_idx += 1
                        
                        c_rh1 = ws.cell(row=row_idx, column=1, value="Data de Vencimento")
                        c_rh2 = ws.cell(row=row_idx, column=2, value="Valor Total")
                        for cell in [c_rh1, c_rh2]:
                            cell.fill = header_fill; cell.font = font_white; cell.alignment = center_align
                        row_idx += 1
                        
                        df_res = df_ativ.groupby(col_venc_ativa)['Valor Total Fatura'].sum().reset_index()
                        df_res['D_Ord'] = pd.to_datetime(df_res[col_venc_ativa], format='%d/%m/%Y', errors='coerce')
                        for _, rs in df_res.sort_values('D_Ord').iterrows():
                            try:
                                c_rv = ws.cell(row=row_idx, column=1, value=pd.to_datetime(rs[col_venc_ativa], format='%d/%m/%Y'))
                                c_rv.number_format = 'DD/MM/YYYY'
                            except:
                                ws.cell(row=row_idx, column=1, value=str(rs[col_venc_ativa]))
                            
                            c_rt = ws.cell(row=row_idx, column=2, value=float(rs['Valor Total Fatura']))
                            c_rt.number_format = 'R$ #,##0.00'
                            row_idx += 1
                            
                        row_idx += 2
                    
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    c_rg = ws.cell(row=row_idx, column=1, value="RESUMO GERAL (TODOS OS SETORES)")
                    c_rg.fill = sector_fill; c_rg.font = font_bold; c_rg.alignment = center_align
                    row_idx += 1
                    
                    c_gh1 = ws.cell(row=row_idx, column=1, value="Data de Vencimento")
                    c_gh2 = ws.cell(row=row_idx, column=2, value="Valor Total")
                    for cell in [c_gh1, c_gh2]:
                        cell.fill = header_fill; cell.font = font_white; cell.alignment = center_align
                    row_idx += 1
                    
                    df_g = df_pendente_envio.groupby(col_venc_ativa)['Valor Total Fatura'].sum().reset_index()
                    df_g['D'] = pd.to_datetime(df_g[col_venc_ativa], format='%d/%m/%Y', errors='coerce')
                    for _, res_g in df_g.sort_values('D').iterrows():
                        try:
                            c_gv = ws.cell(row=row_idx, column=1, value=pd.to_datetime(res_g[col_venc_ativa], format='%d/%m/%Y'))
                            c_gv.number_format = 'DD/MM/YYYY'
                        except:
                            ws.cell(row=row_idx, column=1, value=str(res_g[col_venc_ativa]))
                        
                        c_gt = ws.cell(row=row_idx, column=2, value=float(res_g['Valor Total Fatura']))
                        c_gt.number_format = 'R$ #,##0.00'
                        row_idx += 1

                    from openpyxl.utils import get_column_letter
                    for i, col in enumerate(ws.columns, 1):
                        max_l = 0
                        col_letter = get_column_letter(i)
                        for cell in col:
                            try:
                                if cell.value: max_l = max(max_l, len(str(cell.value)))
                            except: pass
                        ws.column_dimensions[col_letter].width = max_l + 4

                col_btn_gerar, col_vazia1, col_vazia2 = st.columns([1, 2, 2])
                with col_btn_gerar:
                    st.download_button(
                        label="🚀 Gerar Relatório Financeiro",
                        data=buffer.getvalue(),
                        file_name=f"Financeiro_{mes_auditoria.replace('/', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        on_click=finalizar_lote_db,
                        args=(df_pendente_envio, mes_auditoria, col_venc_ativa),
                        use_container_width=True
                    )
            else:
                st.success(f"✅ Não existe pendência de envio para as faturas da CPFL do mês {mes_auditoria}.")

            st.divider()
            with st.expander("📜 Gestão de Envios (Visualizar ou Reverter)"):
                if not df_enviados.empty:
                    df_hist_nomes = pd.merge(df_enviados, df_faturas[['UC', 'Nome da Unidade']].drop_duplicates(), left_on='unidade_consumidora', right_on='UC', how='left')
                    df_hist_nomes['data_envio'] = pd.to_datetime(df_hist_nomes['data_envio']).dt.strftime('%d/%m/%Y %H:%M')
                    st.write("Selecione para **REVERTER** (faturas voltam para a lista acima):")
                    
                    evento_hist = st.dataframe(
                        df_hist_nomes[['id', 'Nome da Unidade', 'unidade_consumidora', 'data_envio', 'valor_fatura']],
                        use_container_width=True, hide_index=True, on_select="rerun", selection_mode="multi-row",
                        column_config={"id": None, "valor_fatura": st.column_config.NumberColumn("Valor (R$)", format="%.2f")},
                        key=f"tabela_reversao_{mes_auditoria}"
                    )
                    
                    if len(evento_hist.selection.rows) > 0:
                        ids_reverter = [int(df_hist_nomes.iloc[i]['id']) for i in evento_hist.selection.rows]
                        if st.button(f"🔄 Reverter {len(ids_reverter)} selecionada(s)", type="secondary"):
                            cursor = conexao.cursor()
                            cursor.execute(f"DELETE FROM historico_financeiro WHERE id IN ({','.join(['%s']*len(ids_reverter))})", tuple(ids_reverter))
                            conexao.commit(); st.rerun()
                else:
                    st.info("Nenhum envio registrado para este mês.")

        # --- SUB-ABA 2: PENDÊNCIAS DE CARGA (LOGICA ATUALIZADA ACL DUAL) ---
        with tab_pendencias:
            linhas_pendentes = []
            hoje = pd.Timestamp.today().normalize()
            
            mes_str, ano_str = mes_auditoria.split('/')
            mes_map_num = {'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6, 
                           'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12}
            mes_num_auditoria = mes_map_num.get(mes_str.upper(), 1)
            ano_num_auditoria = int(ano_str)
            data_auditoria = pd.Timestamp(year=ano_num_auditoria, month=mes_num_auditoria, day=1)

            for _, row in df_cadastro.iterrows():
                uc = str(row['unidade_consumidora']).strip()
                nome = row['nome_unidade']
                dia_cadastrado = int(row['dia_vencimento']) if pd.notna(row['dia_vencimento']) else 10
                
                # Descobre se a UC está no Mercado Livre (Verifica se possui uc_cemig preenchida)
                is_acl = pd.notna(row.get('uc_cemig')) and str(row.get('uc_cemig')).strip() != "" and str(row.get('uc_cemig')).strip().upper() != "NONE"
                
                # Filtra os registros salvos dessa unidade na competência avaliada
                faturas_uc_mes = df_mes[df_mes['UC'] == uc]
                
                # No ACL, separamos as faturas pela coluna 'Classificação'
                tem_cpfl = not faturas_uc_mes[faturas_uc_mes['Classificação'] != 'Mercado Livre - ACL'].empty
                tem_cemig = not faturas_uc_mes[faturas_uc_mes['Classificação'] == 'Mercado Livre - ACL'].empty
                
                # LÓGICA DE AUDITORIA DE CARGA
                status_pendencia = ""
                if is_acl:
                    if not tem_cpfl and not tem_cemig:
                        status_pendencia = "❌ Faltam Ambas (CPFL + CEMIG)"
                    elif not tem_cpfl:
                        status_pendencia = "⚠️ Falta Distribuidora (CPFL)"
                    elif not tem_cemig:
                        status_pendencia = "⚡ Falta Comercializadora (CEMIG)"
                else:
                    if not tem_cpfl:
                        status_pendencia = "❌ Falta Fatura (CPFL)"
                
                # Se faltar qualquer um dos documentos, a UC entra na lista de pendências
                if status_pendencia:
                    faturas_anteriores = df_faturas[(df_faturas['UC'] == uc) & (df_faturas['Data Referência Oculta'] < data_auditoria)]
                    venc_previsto_data = None
                    
                    if not faturas_anteriores.empty:
                        fatura_recente = faturas_anteriores.sort_values('Data Referência Oculta', ascending=False).iloc[0]
                        data_ref_anterior = fatura_recente['Data Referência Oculta']
                        
                        col_v_act = 'Vencimento CPFL' if 'Vencimento CPFL' in fatura_recente else 'Vencimento'
                        venc_anterior_str = fatura_recente[col_v_act]
                        
                        try:
                            venc_anterior_dt = pd.to_datetime(venc_anterior_str, format='%d/%m/%Y')
                            diff_meses = (data_auditoria.year - data_ref_anterior.year) * 12 + (data_auditoria.month - data_ref_anterior.month)
                            
                            novo_mes_venc = venc_anterior_dt.month + diff_meses
                            novo_ano_venc = venc_anterior_dt.year + (novo_mes_venc - 1) // 12
                            novo_mes_venc = ((novo_mes_venc - 1) % 12) + 1
                            
                            ultimo_dia = calendar.monthrange(novo_ano_venc, novo_mes_venc)[1]
                            dia_seguro = min(dia_cadastrado, ultimo_dia)
                            venc_previsto_data = pd.Timestamp(year=novo_ano_venc, month=novo_mes_venc, day=dia_seguro)
                        except:
                            pass
                            
                    if venc_previsto_data is None:
                        novo_mes_venc = mes_num_auditoria + 1
                        novo_ano_venc = ano_num_auditoria
                        if novo_mes_venc > 12:
                            novo_mes_venc = 1
                            novo_ano_venc += 1
                        ultimo_dia = calendar.monthrange(novo_ano_venc, novo_mes_venc)[1]
                        dia_seguro = min(dia_cadastrado, ultimo_dia)
                        venc_previsto_data = pd.Timestamp(year=novo_ano_venc, month=novo_mes_venc, day=dia_seguro)
                    
                    dias_restantes = (venc_previsto_data - hoje).days
                    
                    def semaforo_urgencia(dias):
                        if dias <= 10: return '🔴 Crítico'
                        elif dias <= 20: return '🟠 Atenção'
                        else: return '🟢 No Prazo'
                        
                    linhas_pendentes.append({
                        'Urgência': semaforo_urgencia(dias_restantes),
                        'UC CPFL': uc,
                        'Nome da Unidade': nome,
                        'Ambiente': 'Mercado Livre (ACL)' if is_acl else 'Cativo (ACR)',
                        'Status da Carga': status_pendencia,
                        'Vencimento Previsto': venc_previsto_data.strftime('%d/%m/%Y'),
                        'Dias_Num': dias_restantes
                    })
            
            # Exibição dos dados estruturados
            if linhas_pendentes:
                df_exibir_pendencias = pd.DataFrame(linhas_pendentes)
                df_exibir_pendencias = df_exibir_pendencias.sort_values('Dias_Num', ascending=True).drop(columns=['Dias_Num'])
                
                st.warning(f"🚨 Existem **{len(df_exibir_pendencias)}** unidades com pendências de carga para {mes_auditoria}.")
                
                def colorir_celulas(valor):
                    if 'Crítico' in str(valor):
                        return 'background-color: #FFCDD2; color: #B71C1C; font-weight: bold;'
                    elif 'Atenção' in str(valor):
                        return 'background-color: #FFE0B2; color: #E65100; font-weight: bold;'
                    elif 'No Prazo' in str(valor):
                        return 'background-color: #C8E6C9; color: #1B5E20; font-weight: bold;'
                    return ''
                
                try:
                    tabela_colorida = df_exibir_pendencias.style.map(colorir_celulas, subset=['Urgência'])
                except AttributeError:
                    tabela_colorida = df_exibir_pendencias.style.applymap(colorir_celulas, subset=['Urgência'])
                    
                st.dataframe(tabela_colorida, use_container_width=True, hide_index=True)
            else:
                st.success(f"Excelente! Todas as faturas (CPFL e Comercializadoras ACL) para o mês {mes_auditoria} já foram carregadas no sistema.")

        conexao.close()

# ==========================================
# ABA DADOS
# ==========================================
with aba_dados:
    df = carregar_dados()
    
    if not df.empty:
        st.markdown("##### 📋 Histórico Geral")
        st.markdown("💡 **Dica:** Marque as caixinhas no início das linhas para excluir múltiplos registros de uma só vez.")

        for col in df.columns:
            # Remove apenas o fuso horário das datas. O resto dos dados fica intacto!
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

        # --- NOVO PAINEL DE FILTROS ---
        st.write("") # Dá um espacinho
        
        # Dividimos em 4 colunas agora
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        
        # Cria as caixas de seleção buscando os dados únicos
        filtro_mes = col_f1.multiselect("📅 Filtrar por Referência", options=df['Mês Referência'].unique())
        filtro_classe = col_f2.multiselect("⚡ Filtrar Classificação", options=df['Classificação'].unique())
        filtro_uc = col_f3.multiselect("📍 Filtrar por UC", options=df['UC'].unique())
        filtro_busca = col_f4.text_input("🔍 Busca Livre", placeholder="Nome da unidade...")

        # Aplica a "peneira" sequencial nos dados
        df_filtrado = df
        
        if filtro_mes:
            df_filtrado = df_filtrado[df_filtrado['Mês Referência'].isin(filtro_mes)]
        if filtro_classe:
            df_filtrado = df_filtrado[df_filtrado['Classificação'].isin(filtro_classe)]
        if filtro_uc:
            df_filtrado = df_filtrado[df_filtrado['UC'].isin(filtro_uc)]
        if filtro_busca:
            df_filtrado = df_filtrado[
                df_filtrado['UC'].astype(str).str.contains(filtro_busca, case=False, na=False) |
                df_filtrado['Nome da Unidade'].astype(str).str.contains(filtro_busca, case=False, na=False)
            ]

        # --- TABELA NATIVA (AGORA COM DADOS FILTRADOS) ---
        evento = st.dataframe(
            df_filtrado, # Usa a tabela já filtrada
            hide_index=True,
            use_container_width=True,
            height=400,
            column_config={
                "id": None, 
                "Data Referência Oculta": None,
                "Valor Total Cons. Reat": None,
                "Valor Total Dem. Reat.": None
            },
            selection_mode="multi-row",
            on_select="rerun"
        )
        
        linhas_selecionadas = evento.selection.rows
        
        if len(linhas_selecionadas) > 0:
            # Puxa o ID correto da tabela filtrada
            ids_para_excluir = [int(df_filtrado.iloc[i]['id']) for i in linhas_selecionadas]
            qtd_selecionada = len(ids_para_excluir)
            
            st.markdown(f"🔴 **{qtd_selecionada} Fatura(s) Selecionada(s)** para exclusão.")
            
            if st.session_state.get('confirmar_exclusao_ids') != ids_para_excluir:
                if st.button("🗑️ Excluir Selecionadas"):
                    st.session_state['confirmar_exclusao_ids'] = ids_para_excluir
                    st.rerun()
            else:
                st.warning(f"⚠️ TEM CERTEZA? Isso apagará permanentemente {qtd_selecionada} fatura(s) do banco de dados!")
                col1, col2, _ = st.columns([1, 1, 3]) 
                
                with col1:
                    if st.button("✅ Sim, apagar agora!", type="primary"):
                        conexao = obter_conexao()
                        c = conexao.cursor()
                        
                        # Preparação SQL para o Postgres (%s)
                        placeholders = ','.join('%s' for _ in ids_para_excluir)
                        query_exclusao = f"DELETE FROM faturas_cpfl WHERE id IN ({placeholders})"
                        
                        c.execute(query_exclusao, tuple(ids_para_excluir))
                        conexao.commit()
                        conexao.close()
                        
                        carregar_dados.clear()
                        
                        st.session_state['confirmar_exclusao_ids'] = None
                        st.success(f"{qtd_selecionada} fatura(s) excluída(s) com sucesso!")
                        st.rerun()
                with col2:
                    if st.button("❌ Cancelar"):
                        st.session_state['confirmar_exclusao_ids'] = None
                        st.rerun()

        st.divider()
        @st.cache_data(show_spinner=False, ttl=120, max_entries=1)
        def gerar_excel_cache(df_para_excel):
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_para_excel.to_excel(writer, index=False, sheet_name='Historico_CPFL')
            return buffer.getvalue()

        df_limpo = df.drop(columns=['id', 'Data Referência Oculta'], errors='ignore')
        arquivo_excel_pronto = gerar_excel_cache(df_limpo)
        
        st.download_button(
            label="📥 Exportar Histórico Completo para Excel", 
            data=arquivo_excel_pronto, 
            file_name="Historico_Faturas_Geral.xlsx", 
            type="primary"
        )
    else:
        st.info("Nenhum dado encontrado.")

# ==========================================
# ABA ESPELHO DE FATURA (CORREÇÃO DE ESTADO/CACHE DA TELA)
# ==========================================
with aba_espelho:
    st.markdown("##### 📑 Espelho Técnico e Edição de Fatura")

    if 'msg_sucesso_espelho' in st.session_state:
        st.success(st.session_state['msg_sucesso_espelho'])
        st.balloons() 
        del st.session_state['msg_sucesso_espelho']
    
    df_espelho = carregar_dados()

    if df_espelho.empty:
        st.info("O banco de dados está vazio.")
    else:
        # --- FILTROS DE BUSCA ---
        c_busca1, c_busca2, _ = st.columns([2, 2, 4])
        with c_busca1:
            uc_alvo = st.selectbox("📍 Selecione a UC:", options=sorted(df_espelho['UC'].unique()), key="esp_uc")
        
        meses_uc = df_espelho[df_espelho['UC'] == uc_alvo]['Mês Referência'].unique()
        with c_busca2:
            mes_alvo = st.selectbox("📅 Selecione o Mês:", options=meses_uc, key="esp_mes")

        fatura = df_espelho[(df_espelho['UC'] == uc_alvo) & (df_espelho['Mês Referência'] == mes_alvo)]

        if not fatura.empty:
            f = fatura.iloc[0]
            id_fatura = int(f['id'])
            classe = f['Classificação']
            
            st.divider()
            
            with st.form("form_edicao_espelho"):
                col_info1, col_info2, col_info3 = st.columns(3)
                col_info1.markdown(f"**Unidade:** {f['Nome da Unidade']}")
                col_info2.markdown(f"**Classificação:** {classe}")
                col_info3.markdown(f"**Vencimento:** {f['Vencimento CPFL']}")
                
                ed_cons_p = ed_val_p = ed_cons_fp = ed_val_fp = 0.0
                ed_dc_p = ed_dc_fp = ed_dr_p = ed_val_dr_p = ed_dr_fp = ed_val_dr_fp = 0.0
                ed_di_p = ed_v_di_p = ed_di_fp = ed_v_di_fp = 0.0
                ed_ult_p = ed_v_ult_p = ed_ult_fp = ed_v_ult_fp = 0.0
                ed_reat_p = ed_v_reat_p = ed_reat_fp = ed_v_reat_fp = 0.0
                ed_venc_acl = ""

                # --- LÓGICA TARIFA AZUL ---
                if "Azul" in str(classe):
                    st.subheader("🔹 Detalhamento e Ajuste - Tarifa Azul-A4")
                    tab_cons, tab_ultrap, tab_impostos = st.tabs(["📊 Consumo e Demandas", "⚠️ Ultrapassagem e Reativo", "💰 Impostos e Totais"])
                    
                    with tab_cons:
                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown("**⚡ Consumo Ponta (kWh)**")
                            ed_cons_p = st.number_input("Quantidade Ponta", value=float(f['Consumo Ponta']), format="%.2f", key=f"cp_a_{id_fatura}")
                            ed_val_p = st.number_input("Valor Ponta (R$)", value=float(f['Valor Cons. Ponta TUSD'] + f['Valor Cons. Ponta TE']), format="%.2f", key=f"vp_a_{id_fatura}")
                            st.markdown("---")
                            st.markdown("**📉 Demandas Contratadas (kW)**")
                            ed_dc_p = st.number_input("Contratada Ponta", value=float(f['Dem. Contr. Ponta']), format="%.2f", key=f"dcp_a_{id_fatura}")
                            ed_dc_fp = st.number_input("Contratada F.Ponta", value=float(f['Dem. Contr. F.Ponta']), format="%.2f", key=f"dcfp_a_{id_fatura}")
                        with c2:
                            st.markdown("**⚡ Consumo Fora Ponta (kWh)**")
                            ed_cons_fp = st.number_input("Quantidade F.Ponta", value=float(f['Consumo F.Ponta']), format="%.2f", key=f"cfp_a_{id_fatura}")
                            ed_val_fp = st.number_input("Valor F.Ponta (R$)", value=float(f['Valor Cons. F.Ponta TUSD'] + f['Valor Cons. F.Ponta TE']), format="%.2f", key=f"vfp_a_{id_fatura}")
                            st.markdown("---")
                            st.markdown("**📝 Demandas Registradas e Isentas (kW)**")
                            col_p, col_fp = st.columns(2)
                            with col_p:
                                ed_dr_p = st.number_input("Registrada Ponta", value=float(f['Dem. Reg. Ponta']), format="%.2f", key=f"drp_a_{id_fatura}")
                                ed_val_dr_p = st.number_input("Valor Reg. Ponta", value=float(f['Valor Dem. Ponta']), format="%.2f", key=f"vdrp_a_{id_fatura}")
                                ed_di_p = st.number_input("Isenta Ponta", value=float(f['Dem. Isenta Ponta']), format="%.2f", key=f"dip_a_{id_fatura}")
                                ed_v_di_p = st.number_input("Valor Isenta Ponta", value=float(f['Valor Dem. Isenta Ponta']), format="%.2f", key=f"vdip_a_{id_fatura}")
                            with col_fp:
                                ed_dr_fp = st.number_input("Registrada F.Ponta", value=float(f['Dem. Reg. F.Ponta']), format="%.2f", key=f"drfp_a_{id_fatura}")
                                ed_val_dr_fp = st.number_input("Valor Reg. F.Ponta", value=float(f['Valor Dem. F.Ponta']), format="%.2f", key=f"vdrfp_a_{id_fatura}")
                                ed_di_fp = st.number_input("Isenta F.Ponta", value=float(f['Dem. Isenta F.Ponta']), format="%.2f", key=f"difp_a_{id_fatura}")
                                ed_v_di_fp = st.number_input("Valor Isenta F.Ponta", value=float(f['Valor Dem. Isenta F.Ponta']), format="%.2f", key=f"vdifp_a_{id_fatura}")

                    with tab_ultrap:
                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown("**🚫 Ultrapassagem (kW)**")
                            ed_ult_p = st.number_input("Ultrapassagem Ponta", value=float(f['Dem. Ultrap. Ponta']), format="%.2f", key=f"ultp_a_{id_fatura}")
                            ed_v_ult_p = st.number_input("Valor Ultrap. Ponta", value=float(f['Valor Dem. Ultrap. Ponta']), format="%.2f", key=f"vultp_a_{id_fatura}")
                            ed_ult_fp = st.number_input("Ultrapassagem F.Ponta", value=float(f['Dem. Ultrap. F.Ponta']), format="%.2f", key=f"ultfp_a_{id_fatura}")
                            ed_v_ult_fp = st.number_input("Valor Ultrap. F.Ponta", value=float(f['Valor Dem. Ultrap. F.Ponta']), format="%.2f", key=f"vultfp_a_{id_fatura}")
                        with c2:
                            st.markdown("**⚛️ Reativo Exc.**")
                            ed_reat_p = st.number_input("Reativo Ponta", value=float(f['Dem. Reat. Ponta']), format="%.2f", key=f"reatp_a_{id_fatura}")
                            ed_v_reat_p = st.number_input("Valor Reativo Ponta", value=float(f['Valor Dem. Reat. Ponta']), format="%.2f", key=f"vreatp_a_{id_fatura}")
                            ed_reat_fp = st.number_input("Reativo F.Ponta", value=float(f['Dem. Reat. F.Ponta']), format="%.2f", key=f"reatfp_a_{id_fatura}")
                            ed_v_reat_fp = st.number_input("Valor Reativo F.Ponta", value=float(f['Valor Dem. Reat. F.Ponta']), format="%.2f", key=f"vreatfp_a_{id_fatura}")

                # --- LÓGICA TARIFA VERDE ---
                elif "Verde" in str(classe):
                    st.subheader("🟢 Detalhamento e Ajuste - Tarifa Verde-A4")
                    tab_cons, tab_ultrap, tab_impostos = st.tabs(["📊 Consumo e Demanda", "⚠️ Ultrapassagem e Reativo", "💰 Impostos e Totais"])
                    
                    with tab_cons:
                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown("**⚡ Consumo Ponta (kWh)**")
                            ed_cons_p = st.number_input("Quantidade Ponta", value=float(f['Consumo Ponta']), format="%.2f", key=f"cp_v_{id_fatura}")
                            ed_val_p = st.number_input("Valor Ponta (R$)", value=float(f['Valor Cons. Ponta TUSD'] + f['Valor Cons. Ponta TE']), format="%.2f", key=f"vp_v_{id_fatura}")
                        with c2:
                            st.markdown("**⚡ Consumo Fora Ponta (kWh)**")
                            ed_cons_fp = st.number_input("Quantidade F.Ponta", value=float(f['Consumo F.Ponta']), format="%.2f", key=f"cfp_v_{id_fatura}")
                            ed_val_fp = st.number_input("Valor F.Ponta (R$)", value=float(f['Valor Cons. F.Ponta TUSD'] + f['Valor Cons. F.Ponta TE']), format="%.2f", key=f"vfp_v_{id_fatura}")
                            st.markdown("---")
                            st.markdown("**📉 Demanda Única (kW)**")
                            ed_dc_fp = st.number_input("Contratada Única", value=float(f['Dem. Contr. F.Ponta']), format="%.2f", key=f"dcfp_v_{id_fatura}")
                            ed_dr_fp = st.number_input("Registrada Única", value=float(f['Dem. Reg. F.Ponta']), format="%.2f", key=f"drfp_v_{id_fatura}")
                            ed_val_dr_fp = st.number_input("Valor Reg. Única", value=float(f['Valor Dem. F.Ponta']), format="%.2f", key=f"vdrfp_v_{id_fatura}")
                            ed_di_fp = st.number_input("Isenta Única", value=float(f['Dem. Isenta F.Ponta']), format="%.2f", key=f"difp_v_{id_fatura}")
                            ed_v_di_fp = st.number_input("Valor Isenta Única", value=float(f['Valor Dem. Isenta F.Ponta']), format="%.2f", key=f"vdifp_v_{id_fatura}")

                    with tab_ultrap:
                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown("**🚫 Ultrapassagem Única (kW)**")
                            ed_ult_fp = st.number_input("Ultrapassagem", value=float(f['Dem. Ultrap. F.Ponta']), format="%.2f", key=f"ultfp_v_{id_fatura}")
                            ed_v_ult_fp = st.number_input("Valor Ultrapassagem", value=float(f['Valor Dem. Ultrap. F.Ponta']), format="%.2f", key=f"vultfp_v_{id_fatura}")
                        with c2:
                            st.markdown("**⚛️ Reativo Exc.**")
                            ed_reat_p = st.number_input("Reativo Ponta", value=float(f['Dem. Reat. Ponta']), format="%.2f", key=f"rp_v_{id_fatura}")
                            ed_v_reat_p = st.number_input("Valor Reativo P.", value=float(f['Valor Dem. Reat. Ponta']), format="%.2f", key=f"vrp_v_{id_fatura}")
                            ed_reat_fp = st.number_input("Reativo F.Ponta", value=float(f['Dem. Reat. F.Ponta']), format="%.2f", key=f"rfp_v_{id_fatura}")
                            ed_v_reat_fp = st.number_input("Valor Reativo F.P.", value=float(f['Valor Dem. Reat. F.Ponta']), format="%.2f", key=f"vrfp_v_{id_fatura}")

                # --- LÓGICA CONVENCIONAL B3 ---
                elif "B3" in str(classe) or "Convencional" in str(classe):
                    st.subheader("🏠 Detalhamento e Ajuste - Convencional B3")
                    st.info("💡 Unidade do Grupo B: Sem cobrança de demanda ou reativo.")
                    
                    tab_cons, tab_impostos = st.tabs(["📊 Consumo", "💰 Impostos e Totais"])
                    
                    with tab_cons:
                        c1, _ = st.columns([1, 3])
                        with c1:
                            st.markdown("**⚡ Consumo Ativo (kWh)**")
                            ed_cons_fp = st.number_input("Quantidade Total", value=float(f['Total Consumo']), format="%.2f", key=f"c_b3_{id_fatura}")
                            ed_val_fp = st.number_input("Valor Total do Consumo (R$)", value=float(f['Valor Total Consumo']), format="%.2f", key=f"vc_b3_{id_fatura}")

                else:
                    st.warning("Tipo de tarifa não configurada no espelho.")
                    tab_impostos = st.container()

                # --- BLOCO COMUM PARA TODAS AS TARIFAS: IMPOSTOS ---
                with tab_impostos:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**🏛️ Encargos e Tributos**")
                        ed_cip = st.number_input("CIP (R$)", value=float(f['CIP']), format="%.2f", key=f"cip_{id_fatura}")
                        ed_pis = st.number_input("PIS (R$)", value=float(f['Valor PIS']), format="%.2f", key=f"pis_{id_fatura}")
                        ed_cofins = st.number_input("COFINS (R$)", value=float(f['Valor COFINS']), format="%.2f", key=f"cofins_{id_fatura}")
                        ed_icms = st.number_input("ICMS (R$)", value=float(f['Valor ICMS']), format="%.2f", key=f"icms_{id_fatura}")
                    with c2:
                        st.markdown("**🏁 Resumo Financeiro**")
                        lista_bandeiras = ["VERDE", "AMARELA", "VERMELHA I", "VERMELHA II", "ESCASSEZ HÍDRICA"]
                        bandeira_atual = f['Bandeira'] if f['Bandeira'] in lista_bandeiras else "VERDE"
                        
                        ed_bandeira = st.selectbox("Bandeira", lista_bandeiras, index=lista_bandeiras.index(bandeira_atual), key=f"band_{id_fatura}")
                        ed_val_band = st.number_input("Valor Bandeira (R$)", value=float(f['Adicional Bandeira']), format="%.2f", key=f"valband_{id_fatura}")
                        ed_total = st.number_input("TOTAL DA FATURA (R$)", value=float(f['Valor Total Fatura']), format="%.2f", key=f"total_{id_fatura}")
                        venc_acl_banco = str(f.get('Vencimento ACL', '')) if pd.notna(f.get('Vencimento ACL', '')) else ""
                        ed_venc_acl = st.text_input("Vencimento Comercializadora (ACL)", value=venc_acl_banco, key=f"venc_acl_input_{id_fatura}")

                # --- BOTÃO SALVAR CENTRALIZADO ---
                st.write("")
                _, _, col_btn = st.columns([2, 2, 1])
                with col_btn:
                    if st.form_submit_button("💾 Salvar Alterações", type="primary", use_container_width=True):
                        try:
                            conexao = obter_conexao()
                            cursor = conexao.cursor()
                            
                            sql_update = """
                                UPDATE faturas_cpfl SET 
                                    consumo_ponta=%s, valor_cons_ponta_tusd=%s, valor_cons_ponta_te=0.0,
                                    consumo_fora_ponta=%s, valor_cons_fponta_tusd=%s, valor_cons_fponta_te=0.0,
                                    demanda_contratada_ponta=%s, demanda_contratada_fponta=%s,
                                    demanda_registrada_ponta=%s, valor_dem_ponta=%s,
                                    demanda_registrada_fora_ponta=%s, valor_dem_fponta=%s,
                                    demanda_isenta_ponta=%s, valor_dem_isenta_ponta=%s,
                                    demanda_isenta_fora_ponta=%s, valor_dem_isenta_fponta=%s,
                                    demanda_ultrapassagem_ponta=%s, valor_dem_ultrap_ponta=%s,
                                    demanda_ultrapassagem_fora_ponta=%s, valor_dem_ultrap_fponta=%s,
                                    demanda_reativa_ponta=%s, valor_dem_reativa_ponta=%s,
                                    demanda_reativa_fora_ponta=%s, valor_dem_reativa_fponta=%s,
                                    cip=%s, valor_total_pis=%s, valor_total_cofins=%s, valor_total_icms=%s,
                                    tipo_bandeira=%s, adicional_bandeira=%s, valor_total_fatura=%s,
                                    data_vencimento_acl=%s
                                WHERE id = %s
                            """
                            valores = (ed_cons_p, ed_val_p, ed_cons_fp, ed_val_fp, ed_dc_p, ed_dc_fp,
                                       ed_dr_p, ed_val_dr_p, ed_dr_fp, ed_val_dr_fp,
                                       ed_di_p, ed_v_di_p, ed_di_fp, ed_v_di_fp,
                                       ed_ult_p, ed_v_ult_p, ed_ult_fp, ed_v_ult_fp,
                                       ed_reat_p, ed_v_reat_p, ed_reat_fp, ed_v_reat_fp,
                                       ed_cip, ed_pis, ed_cofins, ed_icms,
                                       ed_bandeira, ed_val_band, ed_total, ed_venc_acl, id_fatura)
                            
                            cursor.execute(sql_update, valores)
                            conexao.commit()
                            conexao.close()
                            
                            st.session_state['msg_sucesso_espelho'] = "✅ Alterações salvas com sucesso! Os dados foram atualizados no banco."
                            carregar_dados.clear() 
                            st.rerun() 
                            
                        except Exception as e:
                            st.error(f"🚨 Erro ao salvar no banco de dados: {e}")
        else:
            st.error("Fatura não encontrada.")

# ==========================================
# ABA IMPORTAÇÃO (PDF E EXCEL)
# ==========================================
with aba_pdf:
    st.markdown("##### 📥 Importação de Faturas")
    
    tab_pdf, tab_excel = st.tabs(["📄 Upload de Faturas (PDF)", "📊 Upload em Lote (excel)"])
    
    with tab_pdf:
        st.markdown("Faça o upload dos arquivos PDF originais da CPFL para extração automática.")
        
        # 1. Cria a chave dinâmica do uploader na memória
        if "pdf_uploader_key" not in st.session_state:
            st.session_state["pdf_uploader_key"] = 0

        # 2. Exibe o relatório ANTES do uploader (após o recarregamento da página)
        if "relatorio_pdf" in st.session_state:
            st.markdown("### 📊 Relatório de Processamento")
            rel = st.session_state["relatorio_pdf"]
            if rel["sucessos"] > 0:
                st.success(f"✅ **{rel['sucessos']}** faturas extraídas e salvas com sucesso!")
                st.balloons()
            if rel["duplicadas"] > 0:
                st.warning(f"⚠️ **{rel['duplicadas']}** faturas foram ignoradas pois já existiam no banco de dados.")
            if rel["erros"] > 0:
                st.error(f"❌ **{rel['erros']}** faturas apresentaram erro durante a leitura.")
            
            # Limpa o relatório da memória para não ficar aparecendo para sempre
            del st.session_state["relatorio_pdf"]

        # 3. Adiciona a "key" com a variável dinâmica no componente de upload
        arquivos_upload = st.file_uploader(
            "Selecione as faturas em PDF", 
            type=["pdf"], 
            accept_multiple_files=True,
            key=f"uploader_{st.session_state['pdf_uploader_key']}"
        )
        
        if arquivos_upload:
            if len(arquivos_upload) > 30:
                st.error(f"⚠️ Limite excedido! Você selecionou {len(arquivos_upload)} arquivos.")
                st.warning("💡 Para evitar sobrecarga e travamento do sistema, por favor, envie no máximo 30 faturas por vez. Remova alguns arquivos acima para liberar o botão de envio.")
            else:
                if st.button("🚀 Extrair e Salvar Dados", type="primary"):
                    sucessos = 0
                    duplicadas = 0
                    erros = 0
                    
                    conexao = obter_conexao()
                    c = conexao.cursor()
                    
                    barra_progresso = st.progress(0)
                    total_arquivos = len(arquivos_upload)
                    
                    for i, arquivo in enumerate(arquivos_upload):
                        try:
                            # Lê a primeira página para rotear o arquivo
                            with pdfplumber.open(arquivo) as pdf_temp:
                                texto_identificacao = pdf_temp.pages[0].extract_text()
                            
                            # ==========================================
                            # 1. PROCESSAMENTO FATURA CEMIG (ACL)
                            # ==========================================
                            if "CEMIG" in texto_identificacao.upper() or "VAREJISTA" in texto_identificacao.upper():
                                d_cemig = processar_pdf_cemig(arquivo)
                                uc_alvo = d_cemig['unidade_consumidora']
                                mes_alvo = d_cemig['mes_referencia']
                                
                                # Verifica se a linha da CPFL para essa UC e Mês já existe
                                c.execute("SELECT id FROM faturas_cpfl WHERE unidade_consumidora = %s AND mes_referencia = %s", (uc_alvo, mes_alvo))
                                row_existente = c.fetchone()
                                
                                if row_existente:
                                    # ATUALIZA A MESMA LINHA EXISTENTE
                                    id_linha = row_existente[0]
                                    c.execute("""
                                        UPDATE faturas_cpfl SET
                                            data_vencimento_acl = %s,
                                            consumo_energia_acl_kwh = %s,
                                            tarifa_energia_acl = %s,
                                            valor_energia_acl = %s,
                                            valor_total_acl = %s
                                        WHERE id = %s
                                    """, (
                                        d_cemig['data_vencimento_acl'],
                                        d_cemig['consumo_energia_acl_kwh'],
                                        d_cemig['tarifa_energia_acl'],
                                        d_cemig['valor_energia_acl'],
                                        d_cemig['valor_total_acl'],
                                        id_linha
                                    ))
                                    sucessos += 1
                                else:
                                    # Se a fatura da CPFL ainda não foi carregada, cria a linha inicial preenchendo os dados da CEMIG
                                    c.execute("""
                                        INSERT INTO faturas_cpfl (
                                            unidade_consumidora, nome_unidade, atividade, mes_referencia, classificacao,
                                            data_vencimento_acl, consumo_energia_acl_kwh, tarifa_energia_acl, valor_energia_acl, valor_total_acl
                                        ) VALUES (%s, %s, %s, %s, 'Mercado Livre - ACL', %s, %s, %s, %s, %s)
                                    """, (
                                        uc_alvo, d_cemig['nome_unidade'], d_cemig['atividade'], mes_alvo,
                                        d_cemig['data_vencimento_acl'], d_cemig['consumo_energia_acl_kwh'],
                                        d_cemig['tarifa_energia_acl'], d_cemig['valor_energia_acl'], d_cemig['valor_total_acl']
                                    ))
                                    sucessos += 1

                            # ==========================================
                            # 2. PROCESSAMENTO FATURA CPFL (ACL OU ACR)
                            # ==========================================
                            else:
                                is_cpfl_acl = "LIVRE-A4" in texto_identificacao.upper() or "CLIENTE LIVRE" in texto_identificacao.upper()
                                d_cpfl = processar_pdf_cpfl_acl(arquivo) if is_cpfl_acl else processar_pdf(arquivo)
                                
                                uc_alvo = d_cpfl['unidade_consumidora']
                                mes_alvo = d_cpfl['mes_referencia']
                                
                                c.execute("SELECT id, valor_total_fatura FROM faturas_cpfl WHERE unidade_consumidora = %s AND mes_referencia = %s", (uc_alvo, mes_alvo))
                                row_existente = c.fetchone()
                                
                                if row_existente:
                                    id_linha, v_total_fatura = row_existente
                                    # Se já tem valor da CPFL preenchido, trata como duplicada
                                    if v_total_fatura > 0:
                                        duplicadas += 1
                                    else:
                                        # Se a linha foi criada primeiro pela CEMIG, atualiza com os dados da CPFL
                                        colunas_up = [f"{k} = %s" for k in d_cpfl.keys() if k not in ('id', 'consumo_energia_acl_kwh', 'tarifa_energia_acl', 'valor_energia_acl', 'valor_total_acl', 'data_vencimento_acl')]
                                        valores_up = tuple(d_cpfl[k] for k in d_cpfl.keys() if k not in ('id', 'consumo_energia_acl_kwh', 'tarifa_energia_acl', 'valor_energia_acl', 'valor_total_acl', 'data_vencimento_acl'))
                                        query_up = f"UPDATE faturas_cpfl SET {', '.join(colunas_up)} WHERE id = %s"
                                        c.execute(query_up, valores_up + (id_linha,))
                                        sucessos += 1
                                else:
                                    # Inserção normal da CPFL
                                    colunas = ', '.join(d_cpfl.keys())
                                    placeholders = ', '.join(['%s'] * len(d_cpfl))
                                    valores = tuple(d_cpfl.values())
                                    c.execute(f"INSERT INTO faturas_cpfl ({colunas}) VALUES ({placeholders})", valores)
                                    sucessos += 1
                                
                        except Exception as e:
                            erros += 1
                            st.error(f"Erro ao processar o arquivo '{arquivo.name}': {e}")
                        
                        barra_progresso.progress((i + 1) / total_arquivos)
                        gc.collect()
                        
                    conexao.commit()
                    carregar_dados.clear()
                    conexao.close()
                    
                    # 4. Salva os resultados na memória, altera a chave do uploader e recarrega a tela!
                    st.session_state["relatorio_pdf"] = {"sucessos": sucessos, "duplicadas": duplicadas, "erros": erros}
                    st.session_state["pdf_uploader_key"] += 1
                    st.rerun()

    with tab_excel:
        st.markdown("Faça o upload de uma planilha contendo o histórico de faturas antigas.")
        st.info("💡 **Dica:** O Excel deve conter os nomes exatos das colunas do banco de dados (em letras minúsculas).")
        
        arquivo_excel = st.file_uploader("Selecione a planilha Excel (.xlsx)", type=["xlsx"])
        
        if arquivo_excel is not None:
            if st.button("🚀 Processar e Salvar Faturas", type="primary"):
                try:
                    df_excel = pd.read_excel(arquivo_excel)
                    conexao = obter_conexao()
                    c = conexao.cursor()
                    
                    def formatar_data(valor):
                        if pd.isna(valor) or valor == '' or valor is None: return ""
                        if isinstance(valor, (int, float)):
                            if valor == 0: return ""
                            return (pd.to_datetime('1899-12-30') + pd.to_timedelta(valor, 'D')).strftime('%d/%m/%Y')
                        if hasattr(valor, 'strftime'):
                            return valor.strftime('%d/%m/%Y')
                        return str(valor).strip()

                    inseridas = 0
                    duplicadas = 0
                    
                    # Checando colunas no Postgres
                    c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'faturas_cpfl'")
                    colunas_banco = [col[0] for col in c.fetchall() if col[0] not in ('id', 'data_insercao')]
                    
                    colunas_texto = ['classificacao', 'unidade_consumidora', 'nome_unidade', 'atividade', 
                                     'periodo_leitura_inicio', 'periodo_leitura_fim', 'data_proxima_leitura', 
                                     'mes_referencia', 'data_vencimento', 'tipo_bandeira']

                    for index, row in df_excel.iterrows():
                        uc = str(row.get('unidade_consumidora', '')).strip()
                        mes = str(row.get('mes_referencia', '')).strip()
                        
                        if not uc or not mes or uc == 'nan' or mes == 'nan':
                            continue
                        
                        c.execute("SELECT id FROM faturas_cpfl WHERE unidade_consumidora = %s AND mes_referencia = %s", (uc, mes))
                        if c.fetchone():
                            duplicadas += 1
                            continue
                        
                        dados_inserir = {}
                        for col in colunas_banco:
                            valor = row.get(col, None)
                            
                            if col in colunas_texto:
                                if pd.isna(valor) or valor == '' or valor is None or valor == 0.0:
                                    dados_inserir[col] = ""
                                elif col in ['data_vencimento', 'periodo_leitura_inicio', 'periodo_leitura_fim', 'data_proxima_leitura']:
                                    dados_inserir[col] = formatar_data(valor)
                                else:
                                    dados_inserir[col] = str(valor).strip()
                            else:
                                if pd.isna(valor) or valor == '' or valor is None:
                                    dados_inserir[col] = 0.0
                                else:
                                    try:
                                        dados_inserir[col] = float(valor)
                                    except:
                                        dados_inserir[col] = 0.0
                        
                        col_str = ', '.join(dados_inserir.keys())
                        placeholders = ', '.join(['%s'] * len(dados_inserir))
                        valores = tuple(dados_inserir.values())
                        
                        c.execute(f"INSERT INTO faturas_cpfl ({col_str}) VALUES ({placeholders})", valores)
                        inseridas += 1
                        
                    conexao.commit()
                    carregar_dados.clear()
                    conexao.close()
                    
                    st.success(f"✅ Processamento concluído! **{inseridas}** faturas inseridas com sucesso. **{duplicadas}** faturas ignoradas (já existiam no banco).")
                    if inseridas > 0:
                        st.balloons()
                        
                except Exception as e:
                    st.error(f"Erro ao processar a planilha. Verifique se os nomes das colunas estão corretos. Erro detalhado: {e}")

# ==========================================
# ABA CONFIGURAÇÕES DE CADASTRO E TARIFAS
# ==========================================
with aba_config:
    st.markdown("##### ⚙️ Configurações Gerais do Sistema")
    st.markdown("###### 🏢 Cadastro Manual da Unidade Consumidora")
    
    col, _, _, _ = st.columns(4)
    with col:
        uc_busca = st.text_input("Buscar UC (Ex: 40190245)", value="").strip()
    
    if uc_busca:
        conexao = obter_conexao()
        c = conexao.cursor()
        # Modificado para extrair também o dia do vencimento cadastrado
        c.execute("SELECT nome_unidade, atividade, classificacao, demanda_contratada_ponta, demanda_contratada_fponta, status, dia_vencimento FROM cadastro_uc WHERE unidade_consumidora = %s", (uc_busca,))
        dados_uc = c.fetchone()
        conexao.close()
    else:
        dados_uc = None
    
    if dados_uc:
        v_nome, v_ativ, v_class, v_dc_p, v_dc_fp, v_status, v_dia_venc = dados_uc
    else:
        v_nome, v_ativ, v_class, v_dc_p, v_dc_fp, v_status, v_dia_venc = ("", None, None, 0.0, 0.0, None, 10)
        
    with st.form("form_uc"):
        # Inserindo a key amarrada a uc_busca
        nome_input = st.text_input("Nome da Instalação/Unidade", value=v_nome, placeholder="Ex: Poço 15 - Geisel", key=f"nome_{uc_busca}")
        
        # --- LINHA 1: Dividindo as caixas de seleção ---
        col_ativ, col_class, col_status, col_dia = st.columns(4)
        
        with col_ativ:
            lista_atividades = ["Administrativa", "Água", "Esgoto"]
            idx_ativ = lista_atividades.index(v_ativ) if v_ativ in lista_atividades else None
            ativ_input = st.selectbox("Atividade", lista_atividades, index=idx_ativ, placeholder="Selecione...", key=f"ativ_{uc_busca}")
            
        with col_class:
            lista_classes = ["Tarifa Azul-A4", "Tarifa Verde-A4", "Convencional B3"]
            idx_class = lista_classes.index(v_class) if v_class in lista_classes else None
            classif_input = st.selectbox("Classificação", lista_classes, index=idx_class, placeholder="Selecione...", key=f"class_{uc_busca}")

        with col_status:
            lista_status = ["ATIVA", "INATIVA"]
            idx_status = lista_status.index(v_status) if v_status in lista_status else None
            status_input = st.selectbox("Status de Operação", lista_status, index=idx_status, placeholder="Selecione...", key=f"status_{uc_busca}")
            
        with col_dia:
            # O segredo para resolver o seu problema está aqui (key=f"dia_{uc_busca}")
            dia_venc_input = st.number_input("Dia Previsto de Vencimento", min_value=1, max_value=31, step=1, value=int(v_dia_venc) if v_dia_venc else 10, key=f"dia_{uc_busca}")
        
        # --- AVISOS ---
        if classif_input and "Verde" in classif_input:
            st.info("💡 Na **Tarifa Verde-A4**, informe a Demanda Única no campo 'Fora Ponta'. O campo Ponta será desconsiderado no cálculo.")
        elif classif_input and "B3" in classif_input:
            st.info("💡 Na **Convencional B3**, não existe demanda contratada. Pode deixar os campos zerados.")
        
        # --- LINHA 2: Dividindo as demandas numéricas em 2 colunas ---
        col_dem1, col_dem2 = st.columns(2)
        is_b3 = True if classif_input and "B3" in classif_input else False
        
        with col_dem1:
            dc_p = st.number_input("Demanda Contratada Ponta (kW)", value=float(v_dc_p), format="%.2f", disabled=is_b3, key=f"dcp_{uc_busca}")
        with col_dem2:
            dc_fp = st.number_input("Demanda Contratada Fora Ponta (kW)", value=float(v_dc_fp), format="%.2f", disabled=is_b3, key=f"dcfp_{uc_busca}")
        
        st.write("")
        if 'msg_uc' in st.session_state:
            st.success(st.session_state['msg_uc'])
            del st.session_state['msg_uc']
        
        # --- BOTÃO SALVAR ---
        col_btn, _, _ = st.columns([1, 3, 3])
        with col_btn:
            if st.form_submit_button("Salvar Cadastro da UC", type="primary", use_container_width=True):
                if ativ_input is None or classif_input is None or status_input is None or not nome_input:
                    st.warning("⚠️ Por favor, preencha o Nome, Atividade, Classificação e Status antes de salvar.")
                else:
                    if classif_input == "Tarifa Verde-A4":
                        dc_p = 0.0 
                    elif classif_input == "Convencional B3":
                        dc_p = 0.0
                        dc_fp = 0.0
                    
                    conexao = obter_conexao()
                    c = conexao.cursor()
                    
                    c.execute('''
                        INSERT INTO cadastro_uc (unidade_consumidora, nome_unidade, atividade, classificacao, demanda_contratada_ponta, demanda_contratada_fponta, status, dia_vencimento)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (unidade_consumidora) DO UPDATE SET 
                        nome_unidade = EXCLUDED.nome_unidade,
                        atividade = EXCLUDED.atividade,
                        classificacao = EXCLUDED.classificacao,
                        demanda_contratada_ponta = EXCLUDED.demanda_contratada_ponta,
                        demanda_contratada_fponta = EXCLUDED.demanda_contratada_fponta,
                        status = EXCLUDED.status,
                        dia_vencimento = EXCLUDED.dia_vencimento;
                    ''', (uc_busca, nome_input, ativ_input, classif_input, dc_p, dc_fp, status_input, dia_venc_input))
                    
                    c.execute('''
                        UPDATE faturas_cpfl 
                        SET nome_unidade = %s 
                        WHERE unidade_consumidora = %s;
                    ''', (nome_input, uc_busca))
                    
                    conexao.commit()
                    conexao.close()
                    
                    st.session_state['msg_uc'] = "✅ Cadastro da UC realizado com sucesso!"
                    st.rerun()

    st.divider()
    st.markdown("###### 🔄 Correção do Banco de Dados")
    st.info("Use este botão para corrigir o nome, atividade e classe de TODAS as faturas antigas de uma só vez.")
    
    col_btn, _, _ = st.columns([1, 3, 3])
    with col_btn:
        if st.button("🔄 Corrigir Faturas Antigas", type="primary", use_container_width=True):
            try:
                conexao = obter_conexao()
                c = conexao.cursor()
                c.execute('''
                    UPDATE faturas_cpfl
                    SET nome_unidade = cadastro_uc.nome_unidade,
                        atividade = cadastro_uc.atividade,
                        classificacao = cadastro_uc.classificacao
                    FROM cadastro_uc
                    WHERE faturas_cpfl.unidade_consumidora = cadastro_uc.unidade_consumidora;
                ''')
                linhas_afetadas = c.rowcount
                conexao.commit()
                conexao.close()
                carregar_dados.clear()
                st.success(f"✅ Sincronização concluída! {linhas_afetadas} faturas foram atualizadas.")
            except Exception as e:
                st.error(f"Erro ao sincronizar: {e}")
